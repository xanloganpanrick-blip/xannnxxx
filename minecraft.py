import io
import re
import openpyxl
from openpyxl.styles import PatternFill
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, CallbackQueryHandler, filters, ContextTypes

TOKEN = "8974827703:AAGbUg8kud9LY97Hm0zYxzbJ26RJvI9zLBo"

# Хранилище сессий пользователей (база данных в памяти)
USER_SESSIONS = {}

# =====================================================================
# БАЗА ИМЁН И ТРАНСЛИТЕРАЦИЯ
# =====================================================================
MALE_NAMES = set([
    "Александр", "Алексей", "Алексий", "Анатолий", "Андрей", "Антон", "Аркадий", "Артём", "Артемий",
    "Афанасий", "Борис", "Вадим", "Валентин", "Валерий", "Василий", "Вениамин", "Виктор", "Виталий",
    "Владимир", "Владислав", "Вячеслав", "Геннадий", "Георгий", "Глеб", "Григорий", "Даниил", "Данила",
    "Денис", "Дмитрий", "Евгений", "Егор", "Иван", "Игорь", "Илья", "Кирилл", "Константин", "Лев",
    "Леонид", "Максим", "Михаил", "Никита", "Николай", "Олег", "Павел", "Пётр", "Роман", "Руслан",
    "Сергей", "Степан", "Тимур", "Фёдор", "Филипп", "Юрий", "Яков", "Ярослав", "Адам", "Азамат",
    "Аким", "Ананий", "Анисим", "Аполлон", "Аристарх", "Арсений", "Архип", "Аскольд", "Богдан",
    "Болеслав", "Бронислав", "Будимир", "Венедикт", "Виссарион", "Влад", "Власий", "Всеволод",
    "Гавриил", "Гермоген", "Ефим", "Ефрем", "Захар", "Зиновий", "Иларион", "Иннокентий", "Иосиф",
    "Исидор", "Исак", "Кондрат", "Корней", "Лаврентий", "Лазарь", "Леон", "Лука", "Лукьян", "Макар",
    "Марк", "Маркел", "Митрофан", "Моисей", "Назар", "Никифор", "Нил", "Нифонт", "Онуфрий", "Остап",
    "Панкратий", "Парфён", "Прохор", "Савва", "Самсон", "Северин", "Семён", "Сильвестр", "Симон",
    "Созонт", "Соломон", "Спиридон", "Тарас", "Тихон", "Трофим", "Трифон", "Федот", "Феликс",
    "Феодосий", "Феофан", "Фока", "Харитон", "Харлам", "Христофор", "Эдуард", "Эльдар", "Элий",
    "Эраст", "Юлиан", "Юлий", "Тимофей", "Матвей", "Елисей", "Демид", "Савелий", "Мирон", "Платон",
    "Клим", "Ростислав", "Мстислав", "Добромир", "Ратибор", "Радомир", "Святослав", "Изяслав", "Ким",
    "Нестор", "Феодор", "Авраам", "Аполлинарий", "Ардалион", "Арефий", "Гурий", "Дорофей", "Ермил",
    "Ермолай", "Зотик", "Зосим", "Зосима", "Ипатий", "Ипполит", "Калистрат", "Каллистрат", "Капитон",
    "Климент", "Кузьма", "Лавр", "Леонтий", "Логин", "Лонгин", "Мартын", "Мефодий", "Мина", "Минай",
    "Модест", "Назарий", "Нафанаил", "Неофит", "Никандр", "Никанор", "Никодим", "Никон", "Онисим",
    "Пантелей", "Пантелеймон", "Парамон", "Пимен", "Питирим", "Поликарп", "Порфирий", "Потап",
    "Прокл", "Прокофий", "Прокопий", "Сысой", "Терентий", "Тимон", "Тит", "Устин", "Ферапонт",
    "Флор", "Фома", "Аслан", "Байрам", "Батыр", "Булат", "Джамал", "Заур", "Ибрагим", "Камал",
    "Камиль", "Карим", "Магомед", "Мухаммад", "Мурат", "Назим", "Нурлан", "Рамазан", "Рашид",
    "Рустам", "Тагир", "Умар", "Хасан", "Хусейн", "Шамиль", "Энвер", "Юсуп", "Аветик", "Акоп",
    "Арман", "Армен", "Ашот", "Грант", "Гурген", "Карен", "Левон", "Тигран", "Ваган", "Вахтанг",
    "Гиви", "Зураб", "Каха", "Коте", "Леван", "Нодар", "Реваз", "Серго", "Тенгиз", "Давит",
    "Альберт", "Альфред", "Генрих", "Карл", "Рудольф", "Эрнест", "Эрнст", "Эрих", "Арнольд",
    "Бенедикт", "Гуго", "Иероним", "Маврикий", "Орест", "Серафим", "Алмир", "Ренат", "Айдар",
    "Айрат", "Азат", "Ильдар", "Ильнур", "Марат", "Радик", "Ринат", "Тимерхан", "Семён", "Ильич"
])

FEMALE_NAMES = set([
    "Александра", "Алина", "Алиса", "Алла", "Альбина", "Анастасия", "Анжела", "Анна", "Антонина",
    "Валентина", "Валерия", "Варвара", "Вера", "Вероника", "Виктория", "Галина", "Дарья", "Диана",
    "Екатерина", "Елена", "Елизавета", "Жанна", "Зинаида", "Зоя", "Инга", "Инна", "Ирина", "Карина",
    "Кристина", "Ксения", "Лариса", "Лидия", "Людмила", "Маргарита", "Марина", "Мария", "Марьяна",
    "Надежда", "Наталья", "Нина", "Оксана", "Ольга", "Полина", "Светлана", "София", "Тамара", "Татьяна",
    "Ульяна", "Юлия", "Агафья", "Агния", "Агриппина", "Аграфена", "Аксинья", "Акулина", "Алевтина",
    "Анфиса", "Арина", "Василиса", "Глафира", "Домна", "Евдокия", "Евгения", "Евлампия", "Евфросиния",
    "Зиновия", "Клавдия", "Клара", "Конкордия", "Любовь", "Любава", "Лукерья", "Маланья", "Матрёна",
    "Матрона", "Меланья", "Милена", "Мирослава", "Нелли", "Неонила", "Прасковья", "Пелагея", "Пелагия",
    "Рогнеда", "Рада", "Раиса", "Регина", "Серафима", "Снежана", "Степанида", "Стефания", "Таисия",
    "Феврония", "Фёкла", "Феодора", "Феофания", "Фотина", "Харитина", "Христина", "Эмилия", "Эмма",
    "Эра", "Юнона", "Юстина", "Ева", "Майя", "Милана", "Алёна", "Дарина", "Аделина", "Аделаида",
    "Адель", "Агата", "Агнесса", "Азалия", "Аида", "Амалия", "Амина", "Амира", "Ангелина", "Белла",
    "Берта", "Богдана", "Бронислава", "Венера", "Виолетта", "Вита", "Владислава", "Гелена", "Генриетта",
    "Гертруда", "Гликерия", "Гульнара", "Гульнур", "Диляра", "Динара", "Евгениа", "Жасмин", "Злата",
    "Изабелла", "Иоанна", "Камилла", "Каролина", "Клеопатра", "Корнелия", "Лаура", "Лейла", "Лилия",
    "Лина", "Лола", "Лора", "Луиза", "Магдалена", "Мадина", "Мадлен", "Марика", "Мариска", "Мариям",
    "Марта", "Мила", "Мирра", "Муза", "Нарине", "Натали", "Наталия", "Наиля", "Ника", "Николетта",
    "Нона", "Нора", "Нурия", "Офелия", "Радмила", "Рената", "Роза", "Розалия", "Роксана", "Сабрина",
    "Сатеник", "Сафия", "Сима", "Сюзанна", "Тереза", "Тина", "Урсула", "Фаина", "Фариза", "Фатима",
    "Фелиция", "Флора", "Хадиджа", "Эдита", "Эллина", "Элла", "Элона", "Элеонора", "Эльвира",
    "Юлиана", "Ядвига", "Яна", "Янина", "Гузель", "Гульназ", "Зульфия", "Резеда", "Эльмира",
])

PAT_M = ["ович", "евич", "ич"]
PAT_F = ["овна", "евна", "ична", "инична", "ьична"]
SUR_END = ["ов", "ев", "ёв", "ин", "ын", "ский", "цкий", "ской", "цкой", "зкий", "хов", "шев", "ков", "ров", "лова",
           "нова", "ич"]


def transliterate_en_to_ru(text):
    rules = {
        'ch': 'ч', 'sh': 'ш', 'th': 'т', 'kh': 'х', 'zh': 'ж', 'ee': 'и', 'oo': 'у', 'ya': 'я', 'yu': 'ю',
        'a': 'а', 'b': 'б', 'v': 'в', 'g': 'г', 'd': 'д', 'e': 'е', 'z': 'з', 'i': 'и', 'j': 'дж', 'k': 'к',
        'l': 'л', 'm': 'м', 'n': 'н', 'o': 'о', 'p': 'п', 'r': 'р', 's': 'с', 't': 'т', 'u': 'у', 'f': 'ф',
        'h': 'х', 'c': 'к', 'q': 'к', 'w': 'в', 'x': 'кс', 'y': 'и', 'z': 'з'
    }
    text_lower = text.lower()
    for k, v in rules.items():
        if len(k) == 2: text_lower = text_lower.replace(k, v)
    for k, v in rules.items():
        if len(k) == 1: text_lower = text_lower.replace(k, v)
    return text_lower.capitalize()


def convert_text_if_english(word):
    if re.match(r'^[A-Za-z]+$', word):
        return transliterate_en_to_ru(word)
    return word


def norm_key(s):
    return s.lower().replace('ё', 'е').strip()


def is_patronymic(w):
    lo = norm_key(w)
    for e in PAT_M:
        if lo.endswith(e) and len(lo) > len(e) + 3: return True
    for e in PAT_F:
        if lo.endswith(e) and len(lo) > len(e) + 2: return True
    return False


def is_name(w):
    lo = norm_key(w)
    cap = w[0].upper() + lo[1:]
    return cap in MALE_NAMES or cap in FEMALE_NAMES


def parse_fio(raw):
    if not raw or not str(raw).strip():
        return {'normalized': '', 'is_full': False}
    raw_words = str(raw).strip().split()
    processed_words = []
    for w in raw_words:
        cleaned_w = re.sub(r'[^A-Za-zА-ЯЁа-яё\-]', '', w)
        if cleaned_w:
            trans_w = convert_text_if_english(cleaned_w)
            if trans_w:
                processed_words.append(trans_w[0].upper() + trans_w[1:].lower())
    if not processed_words:
        return {'normalized': '', 'is_full': False}
    if len(processed_words) == 3:
        if processed_words[2].lower().endswith('ич') and processed_words[0].lower().endswith('ич'):
            base_pat = processed_words[2][:-4] if processed_words[2].lower().endswith('ович') else processed_words[2][
                :-2]
            if is_name(base_pat) or is_name(processed_words[1]):
                return {'normalized': ' '.join(processed_words), 'is_full': True}
    surnames, names, patronymics, unknowns = [], [], [], []
    for w in processed_words:
        if is_patronymic(w) and not (is_name(w) and w.lower().endswith('ич') and len(patronymics) > 0):
            patronymics.append(w)
        elif is_name(w):
            names.append(w)
        elif any(norm_key(w).endswith(e) for e in SUR_END):
            surnames.append(w)
        else:
            unknowns.append(w)
    for w in unknowns:
        if not surnames:
            surnames.append(w)
        elif not names:
            names.append(w)
        else:
            patronymics.append(w)
    final_words = surnames + names + patronymics
    normalized = ' '.join(final_words)
    is_full = len(surnames) > 0 and len(names) > 0 and len(patronymics) > 0
    return {'normalized': normalized, 'is_full': is_full}


def find_columns(ws):
    fio_col, phone_col = None, None
    fio_variants = ['фио', 'ф.и.о', 'фамилия', 'имя', 'наименование']
    phone_variants = ['телефон', 'номер', 'phone', 'tel']
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(1, c).value or '').lower()
        if any(v in val for v in fio_variants) and not fio_col: fio_col = c
        if any(v in val for v in phone_variants) and not phone_col: phone_col = c
    if not fio_col or not phone_col:
        for c in range(1, ws.max_column + 1):
            for r in range(2, min(ws.max_row + 1, 15)):
                cell_val = str(ws.cell(r, c).value or '')
                if not phone_col and re.search(r'\+?[78]\d{9,10}', cell_val): phone_col = c
                if not fio_col and re.search(r'[А-ЯЁа-яёA-Za-z]{3,}\s+[А-ЯЁа-яёA-Za-z]{3,}', cell_val): fio_col = c
    return fio_col or 1, phone_col or 2


def clean_phone(raw_phone):
    if not raw_phone: return ""
    digits = re.sub(r'\D', '', str(raw_phone))
    if digits.startswith('8') and len(digits) == 11: digits = '7' + digits[1:]
    if len(digits) == 10: digits = '7' + digits
    return digits


def parse_sauron_report(text_content):
    results = {}
    blocks = re.split(r'(?:ЗАПРОС:|🔎 ЗАПРОС:)\s*(\d+)', text_content)
    if len(blocks) > 1:
        for i in range(1, len(blocks), 2):
            phone = clean_phone(blocks[i])
            block_content = blocks[i + 1]
            found_fios = re.findall(
                r'\b[A-Za-zА-ЯЁ][a-zа-яё]+\s+[A-Za-zА-ЯЁ][a-zа-яё]+(?:\s+[A-Za-zА-ЯЁ][a-zа-яё]+)?\b', block_content)
            best_fio = ""
            for fio in found_fios:
                fio_clean = fio.strip()
                if any(x in fio_clean.lower() for x in ["санкт", "петербург", "область", "россия"]): continue
                if len(fio_clean.split()) == 3:
                    best_fio = fio_clean
                    break
                elif len(fio_clean.split()) > len(best_fio.split()):
                    best_fio = fio_clean
            if best_fio and phone: results[phone] = best_fio
    else:
        lines = text_content.split('\n')
        current_phone = None
        for line in lines:
            if "Телефон:" in line or "ЗАПРОС:" in line:
                m = re.search(r'\b(7\d{10}|8\d{10})\b', line)
                if m: current_phone = clean_phone(m.group(1))
            if current_phone and ("ФИО:" in line or "Имя:" in line or "Личности:" in line or "•" in line):
                m_fio = re.search(
                    r'\b([A-Za-zА-ЯЁ][a-zа-яё]+\s+[A-Za-zА-ЯЁ][a-zа-яё]+(?:\s+[A-Za-zА-ЯЁ][a-zа-яё]+)?)\b', line)
                if m_fio:
                    fio_candidate = m_fio.group(1)
                    if len(fio_candidate.split()) == 3:
                        results[current_phone] = fio_candidate
                    elif current_phone not in results:
                        results[current_phone] = fio_candidate
    return results


async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Привет! Отправь мне `.xlsx` таблицу для очистки и проверки ФИО.")


async def handle_doc(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    doc = update.message.document
    if not doc:
        return

    if doc.file_name.lower().endswith(('.xlsx', '.xls')):
        msg = await update.message.reply_text("Читаю таблицу, удаляю Оглы/Кызы...")
        file = await ctx.bot.get_file(doc.file_id)
        file_bytes = bytes(await file.download_as_bytearray())
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        fio_col, phone_col = find_columns(ws)
        stop_words = ['кызы', 'оглы', 'улы', 'оглу']
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            fio_val = str(ws.cell(r, fio_col).value or '').lower()
            if any(sw in fio_val for sw in stop_words): rows_to_delete.append(r)
        for r in sorted(rows_to_delete, reverse=True): ws.delete_rows(r)

        incomplete_phones = []
        for r in range(2, ws.max_row + 1):
            fio_val = str(ws.cell(r, fio_col).value or '')
            phone_val = clean_phone(ws.cell(r, phone_col).value)
            parsed = parse_fio(fio_val)
            ws.cell(r, fio_col).value = parsed['normalized']

            # Если ФИО не полное (нет отчества/пусто) и есть телефон — пишем только телефон
            if not parsed['is_full'] and phone_val:
                incomplete_phones.append(phone_val)

        if not incomplete_phones:
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            await msg.edit_text("Все ФИО в порядке! Модификаций не требуется.")
            await ctx.bot.send_document(chat_id=chat_id, document=out, filename="checked_" + doc.file_name)
            return

        out_temp = io.BytesIO()
        wb.save(out_temp)
        USER_SESSIONS[chat_id] = {
            'file_bytes': out_temp.getvalue(),
            'fio_col': fio_col,
            'phone_col': phone_col,
            'filename': doc.file_name,
            'incomplete_phones': incomplete_phones
        }

        # Генерируем TXT, в котором будут ИСКЛЮЧИТЕЛЬНО номера телефонов
        txt_content = "\n".join(incomplete_phones) + "\n"
        txt_file = io.BytesIO(txt_content.encode('utf-8'))
        txt_file.seek(0)

        await ctx.bot.send_document(chat_id=chat_id, document=txt_file, filename="incomplete_fio.txt",
                                    caption="кинь в саурон\n\nПосле получения результата — отправь TXT файл обратно сюда.")
        await msg.delete()

    elif doc.file_name.lower().endswith('.txt'):
        if chat_id not in USER_SESSIONS:
            await update.message.reply_text("Сначала отправь мне файл таблицы `.xlsx`!")
            return
        msg = await update.message.reply_text("Распознаю данные из Саурона и обновляю таблицу...")
        file = await ctx.bot.get_file(doc.file_id)
        txt_bytes = await file.download_as_bytearray()
        txt_content = txt_bytes.decode('utf-8', errors='ignore')
        sauron_data = parse_sauron_report(txt_content)
        session = USER_SESSIONS[chat_id]
        wb = openpyxl.load_workbook(io.BytesIO(session['file_bytes']))
        ws = wb.active
        fio_col = session['fio_col']
        phone_col = session['phone_col']

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        found_count = 0
        not_found_count = 0

        for r in range(2, ws.max_row + 1):
            phone_val = clean_phone(ws.cell(r, phone_col).value)
            if phone_val in session['incomplete_phones']:
                if phone_val in sauron_data:
                    new_fio = sauron_data[phone_val]
                    parsed = parse_fio(new_fio)
                    ws.cell(r, fio_col).value = parsed['normalized']
                    found_count += 1
                else:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = red_fill
                    not_found_count += 1

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        # Сохраняем итоговую таблицу
        session['final_bytes'] = out.getvalue()
        USER_SESSIONS[chat_id] = session

        await ctx.bot.send_document(
            chat_id=chat_id,
            document=out,
            filename="final_" + session['filename'],
            caption=f"Готово!\nУспешно добавлено ФИО: {found_count}\nНе найдено (подсвечены красным): {not_found_count}"
        )
        
        # Показываем кнопку START DOBIV
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("ЗАПУСТИТЬ ДОБИВ", callback_data="start_dobiv")
        ]])
        await update.message.reply_text(
            "Хотите запустить добив через основного бота?\n"
            "Нажмите кнопку ниже — бот подготовит TXT с номерами для пробива.",
            reply_markup=kb
        )
        
        await msg.delete()


async def handle_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    chat_id = update.effective_chat.id
    
    if query.data == "start_dobiv":
        if chat_id not in USER_SESSIONS:
            await query.answer("Сессия не найдена. Отправьте новый файл.")
            await query.message.edit_text("Сессия истекла. Отправьте новый XLSX файл.")
            return
        
        session = USER_SESSIONS[chat_id]
        
        # Формируем TXT со всеми номерами из таблицы
        wb = openpyxl.load_workbook(io.BytesIO(session['final_bytes']))
        ws = wb.active
        fio_col = session['fio_col']
        phone_col = session['phone_col']
        
        # Собираем строки для добива: номер и ФИО
        dobiv_lines = []
        phones_only = []
        for r in range(2, ws.max_row + 1):
            phone_val = clean_phone(ws.cell(r, phone_col).value)
            fio_val = str(ws.cell(r, fio_col).value or '').strip()
            if phone_val:
                # Формат: номер (без +) и ФИО
                phone_digits = phone_val.replace('+', '').replace('7', '', 1) if phone_val.startswith('+7') else phone_val
                phone_clean = '7' + phone_digits[-10:] if len(phone_digits) >= 10 else phone_digits
                if fio_val and fio_val != 'None':
                    dobiv_lines.append(f"{phone_clean} {fio_val}")
                else:
                    phones_only.append(phone_clean)
        
        # Отправляем TXT с номерами
        txt_content = '\n'.join(dobiv_lines + phones_only) + '\n'
        txt_file = io.BytesIO(txt_content.encode('utf-8'))
        txt_file.seek(0)
        
        await ctx.bot.send_document(
            chat_id=chat_id,
            document=txt_file,
            filename="dobiv_numbers.txt",
            caption="TXT для добива через основного бота.\nКаждая строка: Номер Имя (или просто номер).\n\nЗагрузите этот файл на сайте X в разделе 'Добив по TXT'."
        )
        
        await query.answer("TXT для добива отправлен!")
        await query.message.edit_text("TXT для добива отправлен! Загрузите его на сайте X.")
        del USER_SESSIONS[chat_id]


if __name__ == '__main__':
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler('start', start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_doc))
    app.add_handler(CallbackQueryHandler(handle_callback))
    print("Бот запущен (v2.0 с ДОБИВ).")
    app.run_polling()