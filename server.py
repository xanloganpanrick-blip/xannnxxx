# server.py - X Backend v17.0 (РћРџРўРРњРР—РђР¦РРЇ: Г—10 Р±С‹СЃС‚СЂРµРµ, С„РёР»СЊС‚СЂ РґР°С‚ РїРµСЂРІС‹Р№, РєСЌС€ СЃС‚СЂРѕРє, РїР°СЂР°Р»Р»РµР»СЊРЅС‹Рµ Р±РѕС‚С‹, batch 200)

# РЈСЃС‚Р°РЅРѕРІРєР°: pip install aiohttp telethon openpyxl

# Р—Р°РїСѓСЃРє: python server.py

# РџРѕСЂС‚: 8765
import asyncio
import json
import os
import re
import time
import traceback
import io
import zipfile
from datetime import datetime
from aiohttp import web
import aiohttp
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError, FloodWaitError
from telethon.sessions import StringSession
from telethon.tl.types import DocumentAttributeFilename
from openpyxl import load_workbook, Workbook
# ====================== НАСТРОЙКИ ======================
# Используйте переменные окружения. Для локальной разработки заданы значения по умолчанию.
API_ID = int(os.environ.get("API_ID", "2985935"))
API_HASH = os.environ.get("API_HASH", "a436d51ced3ec96a65d8414eb8e0a92d")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "sk-ca6b6569a9b64d0a908eb16ec3b69ce5")

# Резервный API ключ (можно добавить OpenAI/Anthropic при необходимости)
BACKUP_API_KEY = os.environ.get("BACKUP_API_KEY", None)
if BACKUP_API_KEY in ("None", "", None):
    BACKUP_API_KEY = None
sessions = {}
user_clients = {}
TEMP_DIR = "temp_files"
os.makedirs(TEMP_DIR, exist_ok=True)
pending_confirms = {}
pending_web_confirms = {}
active_probevs = {}
probev_lock = asyncio.Lock()
stop_requested = False
pause_requested = False

# ====================== MIDDLEWARE ======================
@web.middleware

async def log_and_cors(request, handler):
    ts = datetime.now().strftime('%H:%M:%S')
    print(f"[REQ] {ts}  {request.method} {request.path}  from={request.remote}", flush=True)
    if request.method == "OPTIONS":
        response = web.Response(status=204)
    else:
        try:
            response = await handler(request)
        except Exception as e:
            print(f"[REQ] {ts}  ERROR {request.method} {request.path}: {e}", flush=True)
            traceback.print_exc()
            response = web.json_response({"ok": False, "error": str(e)}, status=500)
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    print(f"[REQ] {ts}  -> {response.status}", flush=True)
    return response

# ====================== РЈРўРР›РРўР« ======================

def capitalize_words(text):
    """РљР°Р¶РґРѕРµ СЃР»РѕРІРѕ СЃ РїСЂРѕРїРёСЃРЅРѕР№ Р±СѓРєРІС‹ (РєСЂРѕРјРµ Р·Р°РіРѕР»РѕРІРєРѕРІ)"""
    if not text:
        return text
    words = str(text).split()
    return ' '.join([w.capitalize() for w in words])

def normalize_fio_local(raw):
    """Приводит ФИО к формату 'Каждое Слово С Прописной'"""
    if not raw:
        return ""
    # Unicode-escapes для кириллицы (U+0400-U+04FF), латиницы и дефиса
    cleaned = re.sub(r'[^A-Za-z\u0400-\u04FF\-]', ' ', str(raw))
    words = [w.strip() for w in cleaned.split() if w.strip()]
    if not words:
        return ""
    return ' '.join([w[0].upper() + w[1:].lower() if w else '' for w in words])

def normalize_address_local(raw):
    """Приводит адрес к формату 'Город, Улица, Дом, Квартира' с прописной буквы"""
    if not raw:
        return ""
    s = str(raw).strip().lower()
    # Удаляем лишние префиксы (Unicode-escapes для кириллицы)
    s = re.sub(r'\b(\u043E\u0431\u043B\u0430\u0441\u0442\u044C|\u043E\u0431\u043B|\u043A\u0440\u0430\u0439|\u0440\u0435\u0441\u043F\u0443\u0431\u043B\u0438\u043A\u0430|\u0430\u043E|\u0440\u0430\u0439\u043E\u043D|\u0440-\u043D)\b\.?\s*', '', s)
    s = re.sub(r'\b(\u0433|\u0433\u043E\u0440\u043E\u0434|\u0433\u043E\u0440)\b\.?\s*', '', s)
    s = re.sub(r'\b(\u0443\u043B|\u0443\u043B\u0438\u0446\u0430|\u043F\u0440-\u0442|\u043F\u0440\u043E\u0441\u043F\u0435\u043A\u0442|\u043F\u0435\u0440|\u043F\u0435\u0440\u0435\u0443\u043B\u043E\u043A|\u043F\u0440|\u043F\u0440\u043E\u0435\u0437\u0434|\u0431-\u0440|\u0431\u0443\u043B\u044C\u0432\u0430\u0440|\u043F\u043B|\u043F\u043B\u043E\u0449\u0430\u0434\u044C|\u043D\u0430\u0431|\u043D\u0430\u0431\u0435\u0440\u0435\u0436\u043D\u0430\u044F|\u0448|\u0448\u043E\u0441\u0441\u0435)\b\.?\s*', '', s)
    s = re.sub(r'\b(\u0434|\u0434\u043E\u043C|\u0432\u043B\u0434|\u0432\u043B\u0430\u0434\u0435\u043D\u0438\u0435)\b\.?\s*', '', s)
    s = re.sub(r'\b(\u043A\u0432|\u043A\u0432\u0430\u0440\u0442\u0438\u0440\u0430|\u043A\u0432-\u0440\u0430)\b\.?\s*', '', s)
    s = re.sub(r'\b(\u043A\u043E\u0440\u043F|\u043A\u043E\u0440\u043F\u0443\u0441|\u0441\u0442\u0440|\u0441\u0442\u0440\u043E\u0435\u043D\u0438\u0435)\b\.?\s*', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    s = re.sub(r'\s*,\s*', ', ', s)
    # Каждое слово с прописной буквы
    parts = []
    for p in s.split(','):
        p = p.strip()
        if p:
            words = p.split()
            cap_words = [w[0].upper() + w[1:].lower() if w else '' for w in words]
            parts.append(' '.join(cap_words))
    return ', '.join(parts)

async def normalize_batch_deepseek(items, prompt_type='fio', retry_count=0):
    """РџР°РєРµС‚РЅР°СЏ РЅРѕСЂРјР°Р»РёР·Р°С†РёСЏ С‡РµСЂРµР· DeepSeek СЃ СЂРµР·РµСЂРІРЅРѕР№ РјРѕРґРµР»СЊСЋ"""
    if not items:
        return []
    if len(items) <= 2:
        if prompt_type == 'fio':
            return [normalize_fio_local(f) for f in items]
        else:
            return [normalize_address_local(a) for a in items]
    # РџСЂРѕР±СѓРµРј DeepSeek
    try:
        async with aiohttp.ClientSession() as session:
            url = "https://api.deepseek.com/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                "Content-Type": "application/json"
            }
            if prompt_type == 'fio':
                system_prompt = "РџСЂРёРІРµРґРё РєР°Р¶РґРѕРµ Р¤РРћ Рє С„РѕСЂРјР°С‚Сѓ: Р¤Р°РјРёР»РёСЏ РРјСЏ РћС‚С‡РµСЃС‚РІРѕ (РєР°Р¶РґРѕРµ СЃР»РѕРІРѕ СЃ РїСЂРѕРїРёСЃРЅРѕР№ Р±СѓРєРІС‹, РѕСЃС‚Р°Р»СЊРЅС‹Рµ СЃС‚СЂРѕС‡РЅС‹Рµ). Р’РµСЂРЅРё С‚РѕР»СЊРєРѕ СЃРїРёСЃРѕРє, РїРѕ РѕРґРЅРѕРјСѓ РЅР° СЃС‚СЂРѕРєСѓ, СЃ РЅРѕРјРµСЂР°РјРё."
                items_text = "\n".join([f"{i+1}. {f}" for i, f in enumerate(items)])
            elif prompt_type == 'address':
                system_prompt = "РџСЂРёРІРµРґРё РєР°Р¶РґС‹Р№ Р°РґСЂРµСЃ Рє С„РѕСЂРјР°С‚Сѓ: Р“РѕСЂРѕРґ, РЈР»РёС†Р°, Р”РѕРј, РљРІР°СЂС‚РёСЂР°. РЈР±РµСЂРё 'РѕР±Р»Р°СЃС‚СЊ', 'РіРѕСЂРѕРґ', 'СѓР»РёС†Р°', 'РґРѕРј', 'РєРІР°СЂС‚РёСЂР°'. РџСЂРёРјРµСЂ: 'РњСѓСЂРјР°РЅСЃРє, РЎС‚Р°СЂРѕСЃС‚РёРЅР°, 69, 112'. РљР°Р¶РґРѕРµ СЃР»РѕРІРѕ СЃ РїСЂРѕРїРёСЃРЅРѕР№ Р±СѓРєРІС‹. Р’РµСЂРЅРё С‚РѕР»СЊРєРѕ СЃРїРёСЃРѕРє, РїРѕ РѕРґРЅРѕРјСѓ РЅР° СЃС‚СЂРѕРєСѓ, СЃ РЅРѕРјРµСЂР°РјРё."
                items_text = "\n".join([f"{i+1}. {a}" for i, a in enumerate(items)])
            elif prompt_type == 'find_apartment':
                system_prompt = "Р”Р»СЏ РєР°Р¶РґРѕРіРѕ Р°РґСЂРµСЃР° Р±РµР· РєРІР°СЂС‚РёСЂС‹ РЅР°Р№РґРё РєРІР°СЂС‚РёСЂСѓ РёР· РїСЂРёРјРµСЂРѕРІ. РЎРѕРїРѕСЃС‚Р°РІР»СЏР№ РїРѕ РіРѕСЂРѕРґСѓ, СѓР»РёС†Рµ Рё РґРѕРјСѓ. Р’РµСЂРЅРё Р°РґСЂРµСЃ СЃ РєРІР°СЂС‚РёСЂРѕР№ РІ С„РѕСЂРјР°С‚Рµ: Р“РѕСЂРѕРґ, РЈР»РёС†Р°, Р”РѕРј, РљРІР°СЂС‚РёСЂР°. РљР°Р¶РґРѕРµ СЃР»РѕРІРѕ СЃ РїСЂРѕРїРёСЃРЅРѕР№ Р±СѓРєРІС‹. Р’РµСЂРЅРё С‚РѕР»СЊРєРѕ СЃРїРёСЃРѕРє, РїРѕ РѕРґРЅРѕРјСѓ РЅР° СЃС‚СЂРѕРєСѓ, СЃ РЅРѕРјРµСЂР°РјРё. Р•СЃР»Рё РєРІР°СЂС‚РёСЂР° РЅРµ РЅР°Р№РґРµРЅР°, РѕСЃС‚Р°РІСЊ Р°РґСЂРµСЃ РєР°Рє РµСЃС‚СЊ."
                items_text = "\n".join([f"{i+1}. {a}" for i, a in enumerate(items)])
            else:
                return items
            payload = {
                "model": "deepseek-chat",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"РќРѕСЂРјР°Р»РёР·СѓР№:\n{items_text}"}
                ],
                "temperature": 0.1,
                "max_tokens": max(800, len(items) * 35)
            }
            async with session.post(url, headers=headers, json=payload, timeout=aiohttp.ClientTimeout(total=max(15, len(items) // 5))) as resp:
                data = await resp.json()
                if data.get("choices"):
                    text = data["choices"][0]["message"]["content"].strip()
                    result = []
                    for line in text.split('\n'):
                        line = line.strip()
                        if not line:
                            continue
                        if re.match(r'^\d+\.', line):
                            line = re.sub(r'^\d+\.\s*', '', line)
                        if prompt_type == 'fio':
                            # РџСЂРёРІРѕРґРёРј Рє С„РѕСЂРјР°С‚Сѓ "РљР°Р¶РґРѕРµ РЎР»РѕРІРѕ РЎ РџСЂРѕРїРёСЃРЅРѕР№"
                            result.append(capitalize_words(line))
                        else:
                            result.append(capitalize_words(line))
                    return result
    except Exception as e:
        print(f"[DEEPSEEK] РћС€РёР±РєР°: {e}")
        # Р•СЃР»Рё РµСЃС‚СЊ СЂРµР·РµСЂРІРЅС‹Р№ РєР»СЋС‡ Рё СЌС‚Рѕ РїРµСЂРІР°СЏ РїРѕРїС‹С‚РєР°
        if BACKUP_API_KEY and retry_count == 0:
            print("[DEEPSEEK] РџРµСЂРµРєР»СЋС‡РµРЅРёРµ РЅР° СЂРµР·РµСЂРІРЅСѓСЋ РјРѕРґРµР»СЊ...")
            # Р—РґРµСЃСЊ РјРѕР¶РЅРѕ РґРѕР±Р°РІРёС‚СЊ OpenAI/Anthropic
            pass
    # Fallback: Р»РѕРєР°Р»СЊРЅР°СЏ РЅРѕСЂРјР°Р»РёР·Р°С†РёСЏ
    if prompt_type == 'fio':
        return [normalize_fio_local(f) for f in items]
    elif prompt_type == 'address' or prompt_type == 'find_apartment':
        return [normalize_address_local(a) for a in items]
    return items

def clean_phone(text):
    if not text:
        return ""
    d = re.sub(r'[^0-9]', '', str(text))
    if len(d) >= 10:
        return '+7' + d[-10:]
    return d

def clean_phone_without_plus(text):
    """РћС‡РёС‰Р°РµС‚ С‚РµР»РµС„РѕРЅ Рё РІРѕР·РІСЂР°С‰Р°РµС‚ Р±РµР· +7 (С‚РѕР»СЊРєРѕ С†РёС„СЂС‹)"""
    if not text:
        return ""
    d = re.sub(r'[^0-9]', '', str(text))
    if len(d) >= 10:
        return '7' + d[-10:]
    return d

def clean_snils(text):
    if not text:
        return ""
    d = re.sub(r'[^0-9]', '', str(text))
    return d[:11] if len(d) >= 11 else d

def parse_date(val):
    if not val:
        return ""
    s = str(val).strip()
    if hasattr(val, 'strftime'):
        return val.strftime('%d.%m.%Y')
    m = re.search(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})', s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = '20' + y if int(y) < 30 else '19' + y
        return f"{d.zfill(2)}.{mo.zfill(2)}.{y}"
    return ""

def extract_phones_from_text(text):
    phones = []
    for phone in re.findall(r'(?:\+?79\d{9})', text):
        clean = re.sub(r'[^0-9]', '', phone)
        if len(clean) == 11 and clean.startswith('79'):
            phones.append(clean)
    return list(set(phones))

def extract_inn_from_text(text):
    inn_match = re.search(r'\b\d{12}\b', text)
    if inn_match:
        return inn_match.group()
    return None

def extract_address_from_report(text):
    """РР·РІР»РµРєР°РµС‚ Р°РґСЂРµСЃ РёР· РѕС‚С‡С‘С‚Р° СЃР°СѓСЂРѕРЅР°"""
    patterns = [
        r'РђРґСЂРµСЃ[:\s]+([^\n]+)',
        r'РљРѕРЅС‚Р°РєС‚РЅС‹Р№ Р°РґСЂРµСЃ[:\s]+([^\n]+)',
        r'РђРґСЂРµСЃ СЂРµРіРёСЃС‚СЂР°С†РёРё[:\s]+([^\n]+)',
        r'РђР”Р Р•РЎ[:\s]+([^\n]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None

async def find_apartments_via_deepseek(addresses_without_apt, addresses_with_apt):
    """РСЃРїРѕР»СЊР·СѓРµС‚ DeepSeek API РґР»СЏ СЃРѕРїРѕСЃС‚Р°РІР»РµРЅРёСЏ Р°РґСЂРµСЃРѕРІ Рё РїРѕРёСЃРєР° РєРІР°СЂС‚РёСЂ"""
    if not addresses_without_apt or not addresses_with_apt:
        return {}
    # Р¤РѕСЂРјРёСЂСѓРµРј Р·Р°РїСЂРѕСЃ: СЃРїРёСЃРѕРє СЃ РєРІР°СЂС‚РёСЂР°РјРё РєР°Рє РїСЂРёРјРµСЂС‹, СЃРїРёСЃРѕРє Р±РµР· РєРІР°СЂС‚РёСЂ РґР»СЏ РїРѕРёСЃРєР°
    examples = "\n".join([f"  РЎ РєРІР°СЂС‚РёСЂРѕР№: {a}" for a in addresses_with_apt[:20]])
    to_find = "\n".join([f"{i+1}. {a}" for i, a in enumerate(addresses_without_apt)])
    prompt = f"""Р•СЃС‚СЊ Р°РґСЂРµСЃР° СЃ РєРІР°СЂС‚РёСЂР°РјРё (РїСЂРёРјРµСЂС‹):
{examples}
Р”Р»СЏ РєР°Р¶РґРѕРіРѕ Р°РґСЂРµСЃР° РЅРёР¶Рµ РЅР°Р№РґРё РєРІР°СЂС‚РёСЂСѓ, СЃРѕРїРѕСЃС‚Р°РІРёРІ СЃ РїСЂРёРјРµСЂР°РјРё РїРѕ РіРѕСЂРѕРґСѓ, СѓР»РёС†Рµ Рё РґРѕРјСѓ.
Р•СЃР»Рё С‚РѕС‡РЅРѕРµ СЃРѕРІРїР°РґРµРЅРёРµ РЅРµ РЅР°Р№РґРµРЅРѕ вЂ” РѕСЃС‚Р°РІСЊ Р°РґСЂРµСЃ Р±РµР· РєРІР°СЂС‚РёСЂС‹.
Р¤РѕСЂРјР°С‚ РѕС‚РІРµС‚Р°: РЅРѕРјРµСЂ. Р“РѕСЂРѕРґ, РЈР»РёС†Р°, Р”РѕРј, РљРІР°СЂС‚РёСЂР° (РєР°Р¶РґРѕРµ СЃР»РѕРІРѕ СЃ РїСЂРѕРїРёСЃРЅРѕР№).
РђРґСЂРµСЃР° РґР»СЏ РїРѕРёСЃРєР° РєРІР°СЂС‚РёСЂ:
{to_find}"""
    try:
        async with aiohttp.ClientSession() as session:
            url = "https://api.deepseek.com/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                "Content-Type": "application/json"
            }
            payload = {
                "model": "deepseek-chat",
                "messages": [
                    {"role": "system", "content": "РўС‹ РїРѕРјРѕС‰РЅРёРє РґР»СЏ СЃРѕРїРѕСЃС‚Р°РІР»РµРЅРёСЏ Р°РґСЂРµСЃРѕРІ Рё РїРѕРёСЃРєР° РєРІР°СЂС‚РёСЂ. РћС‚РІРµС‡Р°Р№ СЃС‚СЂРѕРіРѕ РІ С„РѕСЂРјР°С‚Рµ: РЅРѕРјРµСЂ. РђРґСЂРµСЃ СЃ РєРІР°СЂС‚РёСЂРѕР№. РљР°Р¶РґРѕРµ СЃР»РѕРІРѕ СЃ РїСЂРѕРїРёСЃРЅРѕР№ Р±СѓРєРІС‹."},
                    {"role": "user", "content": prompt}
                ],
                "temperature": 0.1,
                "max_tokens": max(2000, len(addresses_without_apt) * 40)
            }
            async with session.post(url, headers=headers, json=payload, timeout=aiohttp.ClientTimeout(total=max(30, len(addresses_without_apt) // 3))) as resp:
                data = await resp.json()
                if data.get("choices"):
                    text = data["choices"][0]["message"]["content"].strip()
                    result = {}
                    for line in text.split('\n'):
                        line = line.strip()
                        if not line:
                            continue
                        # РР·РІР»РµРєР°РµРј РїРѕСЂСЏРґРєРѕРІС‹Р№ РЅРѕРјРµСЂ Рё Р°РґСЂРµСЃ
                        m = re.match(r'^(\d+)[\.\)]\s*(.+)', line)
                        if m:
                            idx = int(m.group(1)) - 1
                            addr_with_apt = m.group(2).strip()
                            if idx < len(addresses_without_apt):
                                # РР·РІР»РµРєР°РµРј РєРІР°СЂС‚РёСЂСѓ
                                apt_match = re.search(r',\s*(\d+)\s*$', addr_with_apt)
                                if apt_match:
                                    apartment = apt_match.group(1)
                                    result[addresses_without_apt[idx]] = apartment
                    return result
    except Exception as e:
        print(f"[DEEPSEEK-APT] РћС€РёР±РєР°: {e}")
    # Fallback: Р»РѕРєР°Р»СЊРЅРѕРµ СЃРѕРїРѕСЃС‚Р°РІР»РµРЅРёРµ
    result = {}
    for addr_wo in addresses_without_apt:
        # РЈРїСЂРѕС‰Р°РµРј Р°РґСЂРµСЃ Р±РµР· РєРІР°СЂС‚РёСЂС‹ РґРѕ "РіРѕСЂРѕРґ, СѓР»РёС†Р°, РґРѕРј"
        addr_key = re.sub(r',?\s*\d*\s*$', '', addr_wo).strip().lower()
        addr_key = addr_key.rstrip(',').strip()
        for addr_w in addresses_with_apt:
            if addr_key in addr_w.lower():
                apt_match = re.search(r',\s*(\d+)\s*$', addr_w.strip())
                if apt_match:
                    result[addr_wo] = apt_match.group(1)
                    break
    return result

# ====================== MERGE / SPLIT ======================

def merge_tables(tables_data):
    if not tables_data:
        return None
    base_headers = list(tables_data[0]['headers'])
    addr_idx = -1
    for i, h in enumerate(base_headers):
        if h.lower() in ['Р°РґСЂРµСЃСЃ', 'Р°РґСЂРµСЃ', 'address']:
            addr_idx = i
            break
    table_num_col = 'N С‚Р°Р±Р»РёС†С‹'
    new_headers = []
    if addr_idx >= 0:
        new_headers = base_headers[:addr_idx] + [table_num_col] + base_headers[addr_idx:]
    else:
        new_headers = [table_num_col] + base_headers
    all_rows = []
    for table_idx, table in enumerate(tables_data, 1):
        for row in table['rows']:
            row_map = {}
            for i, h in enumerate(table['headers']):
                if i < len(row):
                    row_map[h.lower()] = row[i]
            new_row = []
            for h in new_headers:
                if h == table_num_col:
                    new_row.append(str(table_idx))
                else:
                    new_row.append(row_map.get(h.lower(), ''))
            all_rows.append(new_row)
    return {'headers': new_headers, 'rows': all_rows}

def split_by_table_num(headers, rows):
    table_num_idx = -1
    for i, h in enumerate(headers):
        if h == 'N С‚Р°Р±Р»РёС†С‹':
            table_num_idx = i
            break
    if table_num_idx == -1:
        return None
    clean_headers = [h for i, h in enumerate(headers) if i != table_num_idx]
    grouped = {}
    for row in rows:
        if len(row) <= table_num_idx:
            continue
        table_num = str(row[table_num_idx]).strip()
        if not table_num:
            continue
        clean_row = [v for i, v in enumerate(row) if i != table_num_idx]
        if table_num not in grouped:
            grouped[table_num] = []
        grouped[table_num].append(clean_row)
    result = {}
    for num, rows_data in grouped.items():
        result[num] = {'headers': clean_headers, 'rows': rows_data}
    return result

def geo_filter(headers, rows):
    GEO_COLS = ['РќРѕРјРµСЂ', 'РђРґСЂРµСЃСЃ', 'Р¤РРћ', 'Р”Р°С‚Р°', 'РЎРќРР›РЎ']
    idx_map = {}
    for col in GEO_COLS:
        found = -1
        for i, h in enumerate(headers):
            if h.lower() == col.lower():
                found = i
                break
        if found == -1:
            for i, h in enumerate(headers):
                hl = h.lower()
                if col.lower() in hl:
                    if col.lower() == 'РЅРѕРјРµСЂ' and ('РїР°СЃРїРѕСЂС‚' in hl or 'РёРЅРЅ' in hl or 'passport' in hl or 'inn' in hl):
                        continue
                    found = i
                    break
        idx_map[col] = found
    for col, idx in idx_map.items():
        if idx == -1:
            raise ValueError(f'РљРѕР»РѕРЅРєР° РЅРµ РЅР°Р№РґРµРЅР°: {col}')
    out_rows = []
    phones = set()
    for row in rows:
        phone_raw = row[idx_map['РќРѕРјРµСЂ']] if idx_map['РќРѕРјРµСЂ'] < len(row) else ''
        phone_clean = clean_phone(phone_raw)
        phones_in_cell = []
        if phone_clean:
            phones_in_cell.append(phone_clean)
        else:
            for p in extract_phones_from_text(str(phone_raw)):
                phones_in_cell.append(p)
        if not phones_in_cell:
            phones_in_cell = ['']
        for ph in phones_in_cell:
            if ph:
                phones.add(ph)
            out_rows.append([
                ph,
                row[idx_map['РђРґСЂРµСЃСЃ']] if idx_map['РђРґСЂРµСЃСЃ'] < len(row) else '',
                row[idx_map['Р¤РРћ']] if idx_map['Р¤РРћ'] < len(row) else '',
                row[idx_map['Р”Р°С‚Р°']] if idx_map['Р”Р°С‚Р°'] < len(row) else '',
                row[idx_map['РЎРќРР›РЎ']] if idx_map['РЎРќРР›РЎ'] < len(row) else ''
            ])
    return {'headers': GEO_COLS, 'rows': out_rows, 'phones': phones}

# ====================== TELEGRAM CLIENT ======================

async def get_client(ss):
    if ss in user_clients:
        c = user_clients[ss]
        if not c.is_connected():
            await c.connect()
        if await c.is_user_authorized():
            return c
    c = TelegramClient(StringSession(ss), API_ID, API_HASH)
    await c.connect()
    if await c.is_user_authorized():
        user_clients[ss] = c
        return c
    await c.disconnect()
    raise Exception("РЎРµСЃСЃРёСЏ РЅРµРґРµР№СЃС‚РІРёС‚РµР»СЊРЅР°")

# ====================== РџРћР”РўР’Р•Р Р–Р”Р•РќРРЇ Р§Р•Р Р•Р— Р‘РћРўРђ ======================

async def send_confirm_with_buttons(bot_token, chat_id, stage_name, count, confirm_id, topic_id=None):
    global stop_requested
    if stop_requested:
        return False
    text = f"РџРћР”РўР’Р•Р Р”РРўР• РџР РћР‘РР’\n\nР­С‚Р°Рї: {stage_name}\nРЎС‚СЂРѕРє: {count}"
    buttons = [
        [{"text": "РџРћР”РўР’Р•Р Р”РРўР¬", "callback_data": f"confirm_{confirm_id}"}],
        [{"text": "РџР РћРџРЈРЎРўРРўР¬", "callback_data": f"skip_{confirm_id}"}],
        [{"text": "РћРЎРўРђРќРћР’РРўР¬ Р’РЎРЃ", "callback_data": f"stop_{confirm_id}"}],
        [{"text": "Р•Р©РЃ Р РђР—", "callback_data": f"again_{confirm_id}"}]
    ]
    kb = {"inline_keyboard": buttons}
    payload = {"chat_id": int(chat_id), "text": text, "reply_markup": kb}
    if topic_id:
        payload["message_thread_id"] = int(topic_id)
    try:
        async with aiohttp.ClientSession() as s:
            await s.post(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                json=payload
            )
        task = asyncio.ensure_future(poll_updates_with_buttons(bot_token, chat_id, confirm_id, topic_id))
        # Логируем исключения фоновой задачи
        def _on_poll_done(t):
            try:
                exc = t.exception()
                if exc:
                    print(f"[POLL] Фоновая задача завершилась с ошибкой: {exc}")
            except asyncio.CancelledError:
                pass
        task.add_done_callback(_on_poll_done)
        return True
    except Exception as e:
        print(f"[CONFIRM] РћС€РёР±РєР°: {e}")
        return False

async def poll_updates_with_buttons(bot_token, chat_id, confirm_id, topic_id=None):
    global stop_requested
    offset = 0
    print(f"[POLL] РќР°С‡РёРЅР°СЋ РѕРїСЂРѕСЃ РґР»СЏ confirm_id={confirm_id}")
    while True:
        try:
            async with aiohttp.ClientSession() as s:
                async with s.get(
                    f"https://api.telegram.org/bot{bot_token}/getUpdates",
                    params={"offset": offset, "timeout": 30}
                ) as resp:
                    data = await resp.json()
                    if data.get("ok") and data.get("result"):
                        for upd in data["result"]:
                            offset = upd["update_id"] + 1
                            cb = upd.get("callback_query")
                            if not cb:
                                continue
                            msg = cb.get("message", {})
                            msg_chat_id = str(msg.get("chat", {}).get("id"))
                            msg_topic_id = msg.get("message_thread_id")
                            if msg_chat_id != str(chat_id):
                                continue
                            if topic_id and msg_topic_id != int(topic_id):
                                continue
                            cb_data = cb.get("data", "")
                            await s.post(
                                f"https://api.telegram.org/bot{bot_token}/answerCallbackQuery",
                                json={"callback_query_id": cb["id"]}
                            )
                            msg_id = msg.get("message_id")
                            payload = {"chat_id": int(chat_id), "message_id": msg_id}
                            if topic_id:
                                payload["message_thread_id"] = int(topic_id)
                            if cb_data == f"confirm_{confirm_id}":
                                payload["text"] = "РџРћР”РўР’Р•Р Р–Р”Р•РќРћ! Р’С‹РїРѕР»РЅСЏСЋ..."
                                await s.post(
                                    f"https://api.telegram.org/bot{bot_token}/editMessageText",
                                    json=payload
                                )
                                pending_confirms[confirm_id] = "confirm"
                                print(f"[POLL] РџРћР”РўР’Р•Р Р–Р”Р•РќРћ")
                                return
                            elif cb_data == f"skip_{confirm_id}":
                                payload["text"] = "Р­РўРђРџ РџР РћРџРЈР©Р•Рќ"
                                await s.post(
                                    f"https://api.telegram.org/bot{bot_token}/editMessageText",
                                    json=payload
                                )
                                pending_confirms[confirm_id] = "skip"
                                print(f"[POLL] РџР РћРџРЈР©Р•Рќ")
                                return
                            elif cb_data == f"stop_{confirm_id}":
                                payload["text"] = "РћРЎРўРђРќРћР’Р›Р•РќРћ! Р—Р°РІРµСЂС€Р°СЋ..."
                                await s.post(
                                    f"https://api.telegram.org/bot{bot_token}/editMessageText",
                                    json=payload
                                )
                                pending_confirms[confirm_id] = "stop"
                                stop_requested = True
                                print(f"[POLL] РћРЎРўРђРќРћР’Р›Р•РќРћ")
                                return
                            elif cb_data == f"again_{confirm_id}":
                                payload["text"] = "Р•Р©РЃ Р РђР—! РћС‚РїСЂР°РІР»СЏСЋ Р·Р°РЅРѕРІРѕ..."
                                await s.post(
                                    f"https://api.telegram.org/bot{bot_token}/editMessageText",
                                    json=payload
                                )
                                pending_confirms[confirm_id] = "again"
                                print(f"[POLL] Р•Р©РЃ Р РђР—")
                                return
        except Exception as e:
            print(f"[POLL] РћС€РёР±РєР°: {e}")
        await asyncio.sleep(1)

async def safe_confirm_with_buttons(bot_token, chat_id, stage_name, count, confirm_id, add_log, topic_id=None):
    global stop_requested
    if stop_requested:
        add_log("[x] РћСЃС‚Р°РЅРѕРІРєР° Р·Р°РїСЂРѕС€РµРЅР°")
        return "stop"
    if not bot_token or not chat_id:
        add_log("[v] Р‘РѕС‚ РЅРµ РЅР°СЃС‚СЂРѕРµРЅ - РїСЂРѕРґРѕР»Р¶Р°СЋ Р°РІС‚РѕРјР°С‚РёС‡РµСЃРєРё")
        return "confirm"
    while True:
        sent = await send_confirm_with_buttons(bot_token, chat_id, stage_name, count, confirm_id, topic_id)
        if not sent:
            add_log("[!] РќРµ СѓРґР°Р»РѕСЃСЊ РѕС‚РїСЂР°РІРёС‚СЊ РІ Р±РѕС‚ - РїРѕРІС‚РѕСЂ С‡РµСЂРµР· 5СЃ...")
            await asyncio.sleep(5)
            continue
        add_log(f"[РћР–РР”РђРќРР•] РћС‚РєСЂРѕР№С‚Рµ Р±РѕС‚ РґР»СЏ: {stage_name}")
        while True:
            if confirm_id in pending_confirms:
                r = pending_confirms.pop(confirm_id)
                if r == "stop":
                    stop_requested = True
                    add_log("[x] РћРЎРўРђРќРћР’РљРђ Р’РЎР•РҐ РџР РћР¦Р•РЎРЎРћР’")
                    return "stop"
                if r == "skip":
                    add_log(f"[v] РџР РћРџРЈР©Р•Рќ: {stage_name}")
                    return "skip"
                if r == "confirm":
                    add_log(f"[v] РџРћР”РўР’Р•Р Р–Р”Р•РќРћ: {stage_name}")
                    return "confirm"
                if r == "again":
                    add_log(f"[v] Р•Р©РЃ Р РђР— - РїРѕРІС‚РѕСЂ РґР»СЏ: {stage_name}")
                    break
            await asyncio.sleep(0.5)

async def send_file_to_bot(bot_token, chat_id, filepath, caption="", topic_id=None):
    try:
        async with aiohttp.ClientSession() as s:
            with open(filepath, 'rb') as f:
                data = aiohttp.FormData()
                data.add_field('chat_id', str(chat_id))
                data.add_field('caption', caption)
                data.add_field('document', f)
                if topic_id:
                    data.add_field('message_thread_id', str(topic_id))
                await s.post(f"https://api.telegram.org/bot{bot_token}/sendDocument", data=data)
            print(f"[BOT] Файл отправлен: {caption}")
    except Exception as e:
        print(f"[BOT] Ошибка: {e}")

async def send_zip_to_bot(bot_token, chat_id, zip_path, caption="", topic_id=None):
    try:
        async with aiohttp.ClientSession() as s:
            with open(zip_path, 'rb') as f:
                data = aiohttp.FormData()
                data.add_field('chat_id', str(chat_id))
                data.add_field('caption', caption)
                data.add_field('document', f)
                if topic_id:
                    data.add_field('message_thread_id', str(topic_id))
                await s.post(f"https://api.telegram.org/bot{bot_token}/sendDocument", data=data)
            print(f"[BOT] ZIP отправлен: {caption}")
    except Exception as e:
        print(f"[BOT] ZIP ошибка: {e}")

async def send_txt_to_bot(bot_token, chat_id, content, filename, caption="", topic_id=None):
    """РћС‚РїСЂР°РІР»СЏРµС‚ TXT С„Р°Р№Р» РІ Р±РѕС‚"""
    try:
        txt_buffer = io.BytesIO(content.encode('utf-8'))
        txt_buffer.seek(0)
        async with aiohttp.ClientSession() as s:
            data = aiohttp.FormData()
            data.add_field('chat_id', str(chat_id))
            data.add_field('caption', caption)
            data.add_field('document', txt_buffer, filename=filename)
            if topic_id:
                data.add_field('message_thread_id', str(topic_id))
            await s.post(f"https://api.telegram.org/bot{bot_token}/sendDocument", data=data)
            print(f"[BOT] TXT РѕС‚РїСЂР°РІР»РµРЅ: {filename}")
    except Exception as e:
        print(f"[BOT] TXT РѕС€РёР±РєР°: {e}")

# ====================== Р РђР‘РћРўРђ РЎ Р‘РћРўРђРњР РџР РћР‘РР’Рђ ======================

async def clear_bot(client, bot):
    try:
        e = await client.get_entity(bot)
        await client.send_message(e, "/start")
        await asyncio.sleep(2)
        print(f"[BOT] /start РѕС‚РїСЂР°РІР»РµРЅ РІ {bot}")
    except Exception as ex:
        print(f"[BOT] РћС€РёР±РєР°: {ex}")

async def click_btn(client, bot, text, retries=3):
    e = await client.get_entity(bot)
    for attempt in range(retries):
        if attempt > 0:
            await asyncio.sleep(3)
        try:
            async for msg in client.iter_messages(e, limit=30):
                if msg.buttons and time.time() - msg.date.timestamp() < 300:
                    for row in msg.buttons:
                        for btn in row:
                            if btn.text and text.lower() in btn.text.lower():
                                await btn.click()
                                await asyncio.sleep(2)
                                print(f"[BOT] РќР°Р¶Р°С‚Р° РєРЅРѕРїРєР° '{btn.text}' РІ {bot}")
                                return True
            print(f"[BOT] РљРЅРѕРїРєР° '{text}' РЅРµ РЅР°Р№РґРµРЅР° РІ {bot} (РїРѕРїС‹С‚РєР° {attempt+1})")
        except Exception as ex:
            print(f"[BOT] РћС€РёР±РєР°: {ex}")
    return False

async def wait_xlsx(client, bot, timeout=180, since_msg_id=None):
    e = await client.get_entity(bot)
    start = time.time()
    print(f"[BOT] РћР¶РёРґР°СЋ XLSX РѕС‚ {bot}...")
    while time.time() - start < timeout:
        msgs = await client.get_messages(e, limit=5)
        for msg in msgs:
            if not msg or not msg.document:
                continue
            if since_msg_id is not None and msg.id <= since_msg_id:
                continue
            for a in msg.document.attributes:
                if isinstance(a, DocumentAttributeFilename) and a.file_name.endswith('.xlsx'):
                    print(f"[BOT] РџРѕР»СѓС‡РµРЅ XLSX: {a.file_name}")
                    return msg
        await asyncio.sleep(3)
    print(f"[BOT] XLSX РЅРµ РїРѕР»СѓС‡РµРЅ")
    return None

async def wait_report(client, bot, phone, timeout=300):
    """РћР¶РёРґР°РµС‚ РѕС‚С‡С‘С‚ РѕС‚ Р±РѕС‚Р° РїРѕ РЅРѕРјРµСЂСѓ С‚РµР»РµС„РѕРЅР°"""
    e = await client.get_entity(bot)
    start = time.time()
    print(f"[BOT] РћР¶РёРґР°СЋ РѕС‚С‡С‘С‚ РїРѕ РЅРѕРјРµСЂСѓ {phone} РѕС‚ {bot}...")
    # РћС‚РїСЂР°РІР»СЏРµРј Р·Р°РїСЂРѕСЃ
    await client.send_message(e, phone)
    while time.time() - start < timeout:
        msgs = await client.get_messages(e, limit=10)
        for msg in msgs:
            if not msg or not msg.text:
                continue
            # РџСЂРѕРІРµСЂСЏРµРј, С‡С‚Рѕ СЌС‚Рѕ РѕС‚С‡С‘С‚ РїРѕ РЅР°С€РµРјСѓ РЅРѕРјРµСЂСѓ
            if phone in msg.text and ("РћРўР§Р•Рў" in msg.text or "Р—РђРџР РћРЎ" in msg.text):
                print(f"[BOT] РџРѕР»СѓС‡РµРЅ РѕС‚С‡С‘С‚ РїРѕ РЅРѕРјРµСЂСѓ {phone}")
                return msg.text
        await asyncio.sleep(5)
    print(f"[BOT] РћС‚С‡С‘С‚ РїРѕ РЅРѕРјРµСЂСѓ {phone} РЅРµ РїРѕР»СѓС‡РµРЅ")
    return None

# ====================== Р”РћР‘РР’ РџРћ РќРћРњР•Р РђРњ (РЎРђРЈР РћРќ) ======================

async def dobiv_by_numbers(client, bot_entity, phone, add_log):
    """РџСЂРѕР±РёРІР°РµС‚ РћР”РРќ РЅРѕРјРµСЂ С‡РµСЂРµР· Р±РѕС‚Р° вЂ” РїРѕСЃР»РµРґРѕРІР°С‚РµР»СЊРЅРѕ, РєР°Рє РІ СЌС‚Р°Р»РѕРЅРЅРѕРј РєРѕРґРµ.
    РћС‚РїСЂР°РІР»СЏРµС‚ РЅРѕРјРµСЂ, Р¶РґС‘С‚ РѕС‚РІРµС‚Р° СЃ Р°РґСЂРµСЃРѕРј."""
    try:
        await client.send_message(bot_entity, phone)
        add_log(f"[РЎРђРЈР РћРќ] Р—Р°РїСЂРѕСЃ РЅРѕРјРµСЂР°: {phone}")
        await asyncio.sleep(6)
        # РС‰РµРј РѕС‚РІРµС‚ Р±РѕС‚Р° СЃ Р°РґСЂРµСЃРѕРј
        async for msg in client.iter_messages(bot_entity, limit=15):
            if not msg or not msg.text:
                continue
            # РџСЂРѕРІРµСЂСЏРµРј С‡С‚Рѕ СЃРѕРѕР±С‰РµРЅРёРµ СЃРѕРґРµСЂР¶РёС‚ РѕС‚С‡С‘С‚ РїРѕ РЅР°С€РµРјСѓ РЅРѕРјРµСЂСѓ
            if phone in msg.text and ("РћРўР§Р•Рў" in msg.text.upper() or "Р—РђРџР РћРЎ" in msg.text.upper() or "РђР”Р Р•РЎ" in msg.text.upper()):
                add_log(f"[РЎРђРЈР РћРќ] РћС‚РІРµС‚ РїРѕР»СѓС‡РµРЅ РґР»СЏ {phone}")
                return msg.text
            # РђР»СЊС‚РµСЂРЅР°С‚РёРІРЅРѕ вЂ” РёС‰РµРј Р»СЋР±РѕР№ РѕС‚РІРµС‚ СЃ Р°РґСЂРµСЃРѕРј
            if "РђР”Р Р•РЎ" in msg.text.upper() and len(msg.text) > 50:
                return msg.text
        add_log(f"[РЎРђРЈР РћРќ] РќРµС‚ РѕС‚РІРµС‚Р° РґР»СЏ {phone}")
        return None
    except FloodWaitError as e:
        add_log(f"[РЎРђРЈР РћРќ] FloodWait {e.seconds}СЃ РґР»СЏ {phone}")
        await asyncio.sleep(e.seconds)
        return None
    except Exception as ex:
        add_log(f"[РЎРђРЈР РћРќ] РћС€РёР±РєР° {phone}: {ex}")
        return None

async def dobiv_fio_date_line_by_line(client, bot_entity, fio_date_pairs, add_log):
    """Этап 6: Отправляет ФИО+дата ПОСТРОЧНО (по 1 на строку) в TXT файле боту2.
    Возвращает словарь {phone: (fio, date)} из ответного XLSX."""
    results = {}
    if not fio_date_pairs:
        return results
    # РЎРѕР·РґР°С‘Рј TXT: РєР°Р¶РґР°СЏ СЃС‚СЂРѕРєР° вЂ” "Р¤РРћ Р”Р°С‚Р°"
    txt_lines = []
    for fio, date in fio_date_pairs:
        txt_lines.append(f"{fio} {date}")
    txt_content = "\n".join(txt_lines) + "\n"
    txt_filename = f"s6_{int(time.time())}.txt"
    txt_path = os.path.join(TEMP_DIR, txt_filename)
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(txt_content)
    add_log(f"[Р­РўРђРџ 6] TXT СЃРѕР·РґР°РЅ: {len(fio_date_pairs)} СЃС‚СЂРѕРє")
    # РћС‚РїСЂР°РІР»СЏРµРј TXT Р±РѕС‚Сѓ2
    try:
        await client.send_file(bot_entity, txt_path, caption="Р¤РРћ+Р”Р°С‚Р° (РїРѕСЃС‚СЂРѕС‡РЅРѕ)")
        add_log(f"[Р­РўРђРџ 6] TXT РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚2, РѕР¶РёРґР°РЅРёРµ XLSX...")
    except Exception as e:
        add_log(f"[Р­РўРђРџ 6] РћС€РёР±РєР° РѕС‚РїСЂР°РІРєРё: {e}")
        return results
    # Ждём XLSX ответ
    bot_identifier = getattr(bot_entity, 'username', None) or str(getattr(bot_entity, 'id', 'unknown'))
    msg = await wait_xlsx(client, bot_identifier, timeout=300)
    if not msg:
        add_log("[ЭТАП 6] XLSX не получен от бота2")
        try:
            os.remove(txt_path)
        except Exception:
            pass
        return results
    rpath = os.path.join(TEMP_DIR, f"r6_{int(time.time())}.xlsx")
    await client.download_media(msg, file=rpath)
    # Очищаем TXT файл
    try:
        os.remove(txt_path)
    except Exception:
        pass
    # РџР°СЂСЃРёРј XLSX вЂ” РёС‰РµРј РЅРѕРјРµСЂР° С‚РµР»РµС„РѕРЅРѕРІ
    try:
        wb = load_workbook(rpath, data_only=True)
        ws = wb.active
        # РћРїСЂРµРґРµР»СЏРµРј РєРѕР»РѕРЅРєРё
        fio_col, phone_col = None, None
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=1, column=c).value or "").lower().strip()
            if any(k in val for k in ['С„РёРѕ', 'fio', 'РёРјСЏ', 'С„Р°РјРёР»РёСЏ', 'name']):
                fio_col = c
            if any(k in val for k in ['С‚РµР»РµС„РѕРЅ', 'phone', 'РЅРѕРјРµСЂ', 'С‚РµР»']):
                phone_col = c
        for row in range(2, ws.max_row + 1):
            fio_val = str(ws.cell(row=row, column=fio_col or 1).value or "").strip()
            phone_val = str(ws.cell(row=row, column=phone_col or 2).value or "").strip()
            phone_clean = clean_phone_without_plus(phone_val)
            if phone_clean and len(phone_clean) >= 10 and fio_val:
                # РС‰РµРј СЃРѕРѕС‚РІРµС‚СЃС‚РІСѓСЋС‰СѓСЋ РїР°СЂСѓ Р¤РРћ+РґР°С‚Р° РёР· РёСЃС…РѕРґРЅРѕРіРѕ СЃРїРёСЃРєР°
                for orig_fio, orig_date in fio_date_pairs:
                    if normalize_fio_local(orig_fio).lower() in normalize_fio_local(fio_val).lower():
                        results[phone_clean] = (orig_fio, orig_date)
                        break
        add_log(f"[Р­РўРђРџ 6] РР·РІР»РµС‡РµРЅРѕ РЅРѕРјРµСЂРѕРІ РёР· XLSX: {len(results)}")
    except Exception as e:
        add_log(f"[Р­РўРђРџ 6] РћС€РёР±РєР° РїР°СЂСЃРёРЅРіР° XLSX: {e}")
    return results

async def dobiv_sauron_sequential(client, bot_entity, fio_date_pairs, add_log, ws_ref, cache_ref, wb_ref, result_file_ref):
    """Р­С‚Р°Рї 7: РџРѕСЃР»РµРґРѕРІР°С‚РµР»СЊРЅС‹Р№ РїСЂРѕР±РёРІ С‡РµСЂРµР· РЎР°СѓСЂРѕРЅ вЂ” РєР°Рє РІ СЌС‚Р°Р»РѕРЅРЅРѕРј РєРѕРґРµ.
    РљР°Р¶РґС‹Р№ Р·Р°РїСЂРѕСЃ РѕС‚РїСЂР°РІР»СЏРµС‚СЃСЏ РѕС‚РґРµР»СЊРЅРѕ, Р¶РґС‘Рј РѕС‚РІРµС‚, РёР·РІР»РµРєР°РµРј С‚РµР»РµС„РѕРЅ.
    РЎРЅР°С‡Р°Р»Р° Р¤РРћ+Р”Р°С‚Р°, РµСЃР»Рё РЅРµ РЅР°Р№РґРµРЅРѕ вЂ” РїСЂРѕР±РёРІ РїРѕ РЎРќРР›РЎ/РРќРќ/РџР°СЃРїРѕСЂС‚."""
    filled = 0
    if not fio_date_pairs:
        return filled
    add_log(f"[Р­РўРђРџ 7] РќР°С‡Р°Р»Рѕ РїРѕСЃР»РµРґРѕРІР°С‚РµР»СЊРЅРѕРіРѕ РїСЂРѕР±РёРІР°: {len(fio_date_pairs)} Р·Р°РїСЂРѕСЃРѕРІ")
    for idx, (row_num, fio, date) in enumerate(fio_date_pairs):
        if stop_requested:
            break
        fio_norm = normalize_fio_local(fio)
        query = f"{fio_norm} {date}"
        try:
            # РћС‚РїСЂР°РІР»СЏРµРј Р·Р°РїСЂРѕСЃ
            await client.send_message(bot_entity, query)
            # Р–РґС‘Рј РѕС‚РІРµС‚
            await asyncio.sleep(6)
            # РС‰РµРј РѕС‚РІРµС‚ СЃ С‚РµР»РµС„РѕРЅР°РјРё
            phone_found = None
            async for msg in client.iter_messages(bot_entity, limit=10):
                if not msg or not msg.text:
                    continue
                text_upper = msg.text.upper()
                if "РћРўР§Р•Рў" in text_upper or "РўР•Р›Р•Р¤РћРќР«" in text_upper or "РўР•Р›Р•Р¤РћРќ" in text_upper:
                    phones = extract_phones_from_text(msg.text)
                    if phones:
                        phone_found = phones[0]
                        break
                    # Р•СЃР»Рё РЅРµС‚ РїСЂСЏРјС‹С… С‚РµР»РµС„РѕРЅРѕРІ вЂ” РёС‰РµРј РЎРќРР›РЎ/РРќРќ/РџР°СЃРїРѕСЂС‚ РґР»СЏ РґРѕР±РёРІР°
                    snils_match = re.search(r'\b\d{11}\b', msg.text)
                    inn_match = re.search(r'\b\d{12}\b', msg.text)
                    passport_match = re.search(r'\b\d{10}\b', msg.text)
                    id_value = None
                    id_type = None
                    if snils_match:
                        id_value = snils_match.group()
                        id_type = "РЎРќРР›РЎ"
                    elif inn_match:
                        id_value = inn_match.group()
                        id_type = "РРќРќ"
                    elif passport_match:
                        id_value = passport_match.group()
                        id_type = "РџР°СЃРїРѕСЂС‚"
                    if id_value:
                        add_log(f"[Р­РўРђРџ 7] РЎС‚СЂРѕРєР° {idx+1}: РќР°Р№РґРµРЅ {id_type} {id_value}, РґРѕР±РёРІ...")
                        await client.send_message(bot_entity, id_value)
                        await asyncio.sleep(6)
                        async for msg2 in client.iter_messages(bot_entity, limit=10):
                            if msg2.text:
                                phones2 = extract_phones_from_text(msg2.text)
                                if phones2:
                                    phone_found = phones2[0]
                                    break
                    break
            if phone_found:
                # Р—Р°РїРѕР»РЅСЏРµРј РЅРѕРјРµСЂ РІ С‚Р°Р±Р»РёС†Рµ
                existing = str(ws_ref.cell(row=row_num, column=COL_PHONE).value or "").strip()
                if not existing or existing == 'None' or existing == '0':
                    ws_ref.cell(row=row_num, column=COL_PHONE).value = phone_found
                    filled += 1
                    wb_ref.save(result_file_ref)
                    cache_ref.rebuild()
                    add_log(f"[Р­РўРђРџ 7] РЎС‚СЂРѕРєР° {row_num}: Р—РђРџРћР›РќР•Рќ {phone_found}")
            if (idx + 1) % 5 == 0:
                add_log(f"[Р­РўРђРџ 7] РџСЂРѕРіСЂРµСЃСЃ: {idx + 1}/{len(fio_date_pairs)}, Р·Р°РїРѕР»РЅРµРЅРѕ: {filled}")
        except FloodWaitError as e:
            add_log(f"[Р­РўРђРџ 7] FloodWait {e.seconds}СЃ")
            await asyncio.sleep(e.seconds)
        except Exception as ex:
            add_log(f"[Р­РўРђРџ 7] РћС€РёР±РєР° СЃС‚СЂРѕРєРё {row_num}: {ex}")
            await asyncio.sleep(2)
    add_log(f"[Р­РўРђРџ 7] РРўРћР“Рћ Р·Р°РїРѕР»РЅРµРЅРѕ: {filled}")
    return filled

async def dobiv_apartments_via_txt(client, bot_entity, phones_list, add_log, ws_ref, cache_ref, wb_ref, result_file_ref):
    """Р­С‚Р°Рї 8: РџСЂРѕР±РёРІ РєРІР°СЂС‚РёСЂ С‡РµСЂРµР· TXT. 
    Р’СЃРµ РЅРѕРјРµСЂР° в†’ TXT (1 РЅРѕРјРµСЂ РЅР° СЃС‚СЂРѕРєСѓ) в†’ Р±РѕС‚2 в†’ TXT РѕС‚РІРµС‚ в†’ DeepSeek СЃРѕРїРѕСЃС‚Р°РІР»СЏРµС‚ Р°РґСЂРµСЃР°."""
    if not phones_list:
        add_log("[Р­РўРђРџ 8] РќРµС‚ РЅРѕРјРµСЂРѕРІ РґР»СЏ РїСЂРѕР±РёРІР° РєРІР°СЂС‚РёСЂ")
        return 0
    # РЎРѕР·РґР°С‘Рј TXT СЃ РЅРѕРјРµСЂР°РјРё (1 РЅРѕРјРµСЂ = 1 СЃС‚СЂРѕРєР°)
    unique_phones = list(set([clean_phone_without_plus(p) for p in phones_list if p]))
    txt_content = "\n".join(unique_phones) + "\n"
    txt_path = os.path.join(TEMP_DIR, f"apt_{int(time.time())}.txt")
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(txt_content)
    add_log(f"[Р­РўРђРџ 8] TXT СЃ РЅРѕРјРµСЂР°РјРё СЃРѕР·РґР°РЅ: {len(unique_phones)} РЅРѕРјРµСЂРѕРІ")
    try:
        # РћС‚РїСЂР°РІР»СЏРµРј TXT Р±РѕС‚Сѓ2
        last_msgs_before = await client.get_messages(bot_entity, limit=1)
        last_msg_id = last_msgs_before[0].id if last_msgs_before else 0
        await client.send_file(bot_entity, txt_path, caption=f"РџСЂРѕР±РёРІ РєРІР°СЂС‚РёСЂ: {len(unique_phones)} РЅРѕРјРµСЂРѕРІ")
        add_log("[Р­РўРђРџ 8] TXT РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚2, РѕР¶РёРґР°РЅРёРµ РѕС‚РІРµС‚РЅРѕРіРѕ TXT...")
        # Р–РґС‘Рј TXT РѕС‚РІРµС‚ РѕС‚ Р±РѕС‚Р°2
        await asyncio.sleep(5)
        response_text = None
        for attempt in range(30):  # РґРѕ 5 РјРёРЅСѓС‚
            async for msg in client.iter_messages(bot_entity, limit=5):
                if msg.id <= last_msg_id:
                    continue
                if msg.document:
                    # РџСЂРѕРІРµСЂСЏРµРј TXT РёР»Рё РґРѕРєСѓРјРµРЅС‚
                    for attr in msg.document.attributes:
                        if isinstance(attr, DocumentAttributeFilename):
                            fname = attr.file_name.lower()
                            if fname.endswith('.txt'):
                                # РЎРєР°С‡РёРІР°РµРј TXT
                                rpath = os.path.join(TEMP_DIR, f"apt_resp_{int(time.time())}.txt")
                                await client.download_media(msg, file=rpath)
                                with open(rpath, 'r', encoding='utf-8', errors='ignore') as rf:
                                    response_text = rf.read()
                                add_log(f"[Р­РўРђРџ 8] TXT РѕС‚РІРµС‚ РїРѕР»СѓС‡РµРЅ: {len(response_text)} СЃРёРјРІРѕР»РѕРІ")
                                break
                elif msg.text and len(msg.text) > 200:
                    response_text = msg.text
                    add_log(f"[Р­РўРђРџ 8] РўРµРєСЃС‚РѕРІС‹Р№ РѕС‚РІРµС‚ РїРѕР»СѓС‡РµРЅ: {len(response_text)} СЃРёРјРІРѕР»РѕРІ")
                    break
            if response_text:
                break
            await asyncio.sleep(10)
        if not response_text:
            add_log("[Р­РўРђРџ 8] РћС‚РІРµС‚ РѕС‚ Р±РѕС‚Р°2 РЅРµ РїРѕР»СѓС‡РµРЅ")
            return 0
        # РџР°СЂСЃРёРј РѕС‚РІРµС‚: СЃС‚СЂСѓРєС‚СѓСЂР° СЃ СЂР°Р·РґРµР»РёС‚РµР»СЏРјРё в”Ѓв”Ѓв”Ѓв”Ѓв”Ѓв”Ѓв”Ѓ
        # РљР°Р¶РґС‹Р№ Р±Р»РѕРє РјРµР¶РґСѓ СЂР°Р·РґРµР»РёС‚РµР»СЏРјРё вЂ” РёРЅС„РѕСЂРјР°С†РёСЏ РїРѕ РѕРґРЅРѕРјСѓ РЅРѕРјРµСЂСѓ
        blocks = re.split(r'в”Ѓ{5,}', response_text)
        add_log(f"[Р­РўРђРџ 8] РќР°Р№РґРµРЅРѕ Р±Р»РѕРєРѕРІ: {len(blocks)}")
        # РР·РІР»РµРєР°РµРј Р°РґСЂРµСЃР° РёР· Р±Р»РѕРєРѕРІ
        phone_addresses = {}  # phone -> address
        current_phone = None
        for block in blocks:
            block = block.strip()
            if not block:
                continue
            # РС‰РµРј РЅРѕРјРµСЂ С‚РµР»РµС„РѕРЅР° РІ Р±Р»РѕРєРµ
            phone_match = re.search(r'(?:7)?(\d{10,11})', block)
            if phone_match:
                current_phone = phone_match.group(1)
                if len(current_phone) == 10:
                    current_phone = '7' + current_phone
            # РС‰РµРј Р°РґСЂРµСЃ РІ Р±Р»РѕРєРµ
            addr_match = re.search(r'(?:РђРґСЂРµСЃ|Р°РґСЂРµСЃ|РђР”Р Р•РЎ)[:\s]+([^\n]+)', block)
            if not addr_match:
                # РђР»СЊС‚РµСЂРЅР°С‚РёРІРЅС‹Рµ РїР°С‚С‚РµСЂРЅС‹
                addr_match = re.search(r'(?:Рі\.|СѓР»\.|Рґ\.)\s*[^\n]+', block)
            if current_phone and addr_match:
                address = addr_match.group(1) if addr_match.lastindex else addr_match.group(0)
                address = address.strip()
                phone_addresses[current_phone] = normalize_address_local(address)
        add_log(f"[Р­РўРђРџ 8] РђРґСЂРµСЃРѕРІ РёР·РІР»РµС‡РµРЅРѕ: {len(phone_addresses)}")
        # РЎРѕР±РёСЂР°РµРј Р°РґСЂРµСЃР° СЃ РєРІР°СЂС‚РёСЂР°РјРё Рё Р±РµР·
        addrs_with_apt = []
        addrs_no_apt_items = []
        for r, d in cache_ref.rows.items():
            addr = d['addr']
            if not addr or addr == 'None':
                continue
            norm = normalize_address_local(addr)
            if re.search(r',\s*\d+\s*$', addr):
                if norm not in addrs_with_apt:
                    addrs_with_apt.append(norm)
            else:
                phone = clean_phone(d['phone'])
                if phone and phone != 'None' and phone != '0':
                    addrs_no_apt_items.append({
                        'row': r, 'address': addr,
                        'address_normalized': norm, 'phone': phone
                    })
        # РћР±СЉРµРґРёРЅСЏРµРј РїСЂРёРјРµСЂС‹ РёР· РѕС‚РІРµС‚Р° Рё РёР· С‚Р°Р±Р»РёС†С‹
        all_examples = list(set(addrs_with_apt + list(phone_addresses.values())))
        targets = [item['address_normalized'] for item in addrs_no_apt_items]
        if targets and all_examples:
            add_log(f"[Р­РўРђРџ 8] DeepSeek: {len(targets)} С†РµР»РµР№, {len(all_examples)} РїСЂРёРјРµСЂРѕРІ")
            apt_map = await find_apartments_via_deepseek(targets, all_examples)
            if apt_map:
                filled_count = 0
                for item in addrs_no_apt_items:
                    norm = item['address_normalized']
                    if norm in apt_map:
                        apartment = apt_map[norm]
                        row = item['row']
                        current = str(ws_ref.cell(row=row, column=COL_ADDR).value or "").strip()
                        new_addr = f"{current.rstrip(',')}, {apartment}"
                        ws_ref.cell(row=row, column=COL_ADDR).value = new_addr
                        filled_count += 1
                    elif item['phone'] in phone_addresses:
                        # Р‘РµСЂС‘Рј РїРѕР»РЅС‹Р№ Р°РґСЂРµСЃ РёР· РѕС‚РІРµС‚Р°
                        full_addr = phone_addresses[item['phone']]
                        # РР·РІР»РµРєР°РµРј РєРІР°СЂС‚РёСЂСѓ
                        apt_match = re.search(r',\s*(\d+)\s*$', full_addr)
                        if apt_match:
                            row = item['row']
                            current = str(ws_ref.cell(row=row, column=COL_ADDR).value or "").strip()
                            new_addr = f"{current.rstrip(',')}, {apt_match.group(1)}"
                            ws_ref.cell(row=row, column=COL_ADDR).value = new_addr
                            filled_count += 1
                        else:
                            # Р‘РµСЂС‘Рј РІРµСЃСЊ Р°РґСЂРµСЃ Рё С„РѕСЂРјР°С‚РёСЂСѓРµРј
                            row = item['row']
                            ws_ref.cell(row=row, column=COL_ADDR).value = full_addr
                            filled_count += 1
                wb_ref.save(result_file_ref)
                cache_ref.rebuild()
                add_log(f"[Р­РўРђРџ 8] РРўРћР“Рћ Р·Р°РїРѕР»РЅРµРЅРѕ Р°РґСЂРµСЃРѕРІ/РєРІР°СЂС‚РёСЂ: {filled_count}")
                try:
                    os.remove(txt_path)
                except Exception:
                    pass
                return filled_count
        add_log("[Р­РўРђРџ 8] РЎРѕРІРїР°РґРµРЅРёР№ РЅРµ РЅР°Р№РґРµРЅРѕ")
        try:
            os.remove(txt_path)
        except Exception:
            pass
        return 0
    except Exception as e:
        add_log(f"[Р­РўРђРџ 8] РћС€РёР±РєР°: {e}")
        traceback.print_exc()
        try:
            os.remove(txt_path)
        except Exception:
            pass
        return 0

# ====================== РџРђР РЎРРќР“ XLSX ======================

def parse_xlsx(path):
    res = []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        h = {}
        for col in range(1, ws.max_column + 1):
            v = str(ws.cell(row=1, column=col).value or "").upper().strip()
            if any(k in v for k in ['РРќРќ', 'INN', 'РџРђРЎРџРћР Рў', 'PASSPORT']):
                continue
            if any(k in v for k in ['Р¤РРћ', 'FIO', 'РРњРЇ', 'Р¤РђРњРР›РРЇ', 'NAME']):
                h['fio'] = col
            if any(k in v for k in ['Р”РђРўРђ', 'DATE', 'Р РћР–Р”', 'BIRTH']):
                h['date'] = col
            if any(k in v for k in ['РўР•Р›Р•Р¤РћРќ', 'PHONE', 'РўР•Р›']):
                h['phone'] = col
            elif 'РќРћРњР•Р ' in v and 'РџРђРЎРџРћР Рў' not in v and 'РРќРќ' not in v:
                h['phone'] = col
            if any(k in v for k in ['РЎРќРР›РЎ', 'SNILS']):
                h['snils'] = col
            if any(k in v for k in ['РђР”Р Р•РЎ', 'РђР”Р Р•РЎРЎ', 'ADDR', 'ADDRESS']):
                h['addr'] = col
        for row in range(2, ws.max_row + 1):
            try:
                r = {}
                if 'fio' in h:
                    r['fio'] = normalize_fio_local(ws.cell(row=row, column=h['fio']).value)
                if 'date' in h:
                    r['date'] = parse_date(ws.cell(row=row, column=h['date']).value)
                if 'phone' in h:
                    r['phone'] = clean_phone(ws.cell(row=row, column=h['phone']).value)
                if 'snils' in h:
                    r['snils'] = clean_snils(ws.cell(row=row, column=h['snils']).value)
                if 'addr' in h:
                    r['addr'] = str(ws.cell(row=row, column=h['addr']).value or "").strip()
                if any(v for v in r.values() if v):
                    res.append(r)
            except Exception:
                pass
        return res
    except Exception as e:
        print(f"[PARSE] РћС€РёР±РєР°: {e}")
        return []

# ====================== Р—РђРџРћР›РќР•РќРР• РўРђР‘Р›РР¦Р« ======================
COL_NO = 1
COL_FIO = 2
COL_DATE = 3
COL_PHONE = 4
COL_SNILS = 5
COL_ADDR = 6

def fill_dates_from_response(ws, response_records):
    filled = 0
    print(f"[FILL-DATES] РќР°С‡РёРЅР°СЋ Р·Р°РїРѕР»РЅРµРЅРёРµ РґР°С‚. РћС‚РІРµС‚РѕРІ: {len(response_records)}")
    table_phones = {}
    for row in range(2, ws.max_row + 1):
        phone = str(ws.cell(row=row, column=COL_PHONE).value or "").strip()
        if phone:
            clean = clean_phone(phone)
            if clean:
                table_phones[clean] = row
    for rec in response_records:
        rec_phone = rec.get('phone', '')
        rec_date = rec.get('date', '')
        if not rec_date or not rec_phone:
            continue
        clean_rec = clean_phone(rec_phone)
        if not clean_rec:
            continue
        if clean_rec in table_phones:
            row = table_phones[clean_rec]
            existing = str(ws.cell(row=row, column=COL_DATE).value or "").strip()
            if not existing or existing == 'None':
                ws.cell(row=row, column=COL_DATE).value = rec_date
                filled += 1
                print(f"[FILL-DATES] РЎС‚СЂРѕРєР° {row}: Р—РђРџРћР›РќР•РќРћ")
    print(f"[FILL-DATES] РРўРћР“Рћ Р·Р°РїРѕР»РЅРµРЅРѕ РґР°С‚: {filled}")
    return filled

def fill_phones_from_response(ws, response_records):
    filled = 0
    print(f"[FILL-PHONES] Начинаю заполнение номеров. Ответов: {len(response_records)}")
    table_index = {}
    fio_fallback_index = {}  # fio -> list of row numbers (оптимизация fallback)
    for row in range(2, ws.max_row + 1):
        fio = normalize_fio_local(str(ws.cell(row=row, column=COL_FIO).value or ""))
        date = str(ws.cell(row=row, column=COL_DATE).value or "").strip()
        if fio and date and date != 'None':
            table_index[(fio, date)] = row
        if fio:
            fio_fallback_index.setdefault(fio, []).append(row)
    for rec in response_records:
        rec_fio = normalize_fio_local(rec.get('fio', ''))
        rec_phone = rec.get('phone', '')
        rec_date = rec.get('date', '')
        if not rec_phone or not rec_fio:
            continue
        clean_rec = clean_phone(rec_phone)
        if not clean_rec:
            continue
        key = (rec_fio, rec_date)
        matched_row = None
        if key in table_index:
            matched_row = table_index[key]
        elif rec_fio in fio_fallback_index:
            for row in fio_fallback_index[rec_fio]:
                existing = str(ws.cell(row=row, column=COL_PHONE).value or "").strip()
                if not existing or existing == 'None':
                    matched_row = row
                    break
        if matched_row is not None:
            existing = str(ws.cell(row=matched_row, column=COL_PHONE).value or "").strip()
            if not existing or existing == 'None':
                ws.cell(row=matched_row, column=COL_PHONE).value = clean_rec
                filled += 1
                print(f"[FILL-PHONES] Строка {matched_row}: ЗАПОЛНЕНО")
    print(f"[FILL-PHONES] ИТОГО заполнено номеров: {filled}")
    return filled

def fill_snils_dates(ws, response_records):
    filled = 0
    print(f"[FILL-SNILS] Начинаю заполнение дат по СНИЛС")
    table_snils = {}
    fio_fallback_index = {}  # fio -> list of row numbers (оптимизация fallback)
    for row in range(2, ws.max_row + 1):
        snils = clean_snils(str(ws.cell(row=row, column=COL_SNILS).value or ""))
        if snils and len(snils) >= 11:
            table_snils[snils] = row
        fio = normalize_fio_local(str(ws.cell(row=row, column=COL_FIO).value or ""))
        if fio:
            fio_fallback_index.setdefault(fio, []).append(row)
    for rec in response_records:
        rec_snils = clean_snils(rec.get('snils', ''))
        rec_fio = normalize_fio_local(rec.get('fio', ''))
        rec_date = rec.get('date', '')
        if not rec_date:
            continue
        found = False
        if rec_snils and len(rec_snils) >= 11 and rec_snils in table_snils:
            row = table_snils[rec_snils]
            existing = str(ws.cell(row=row, column=COL_DATE).value or "").strip()
            if not existing or existing == 'None':
                ws.cell(row=row, column=COL_DATE).value = rec_date
                filled += 1
                print(f"[FILL-SNILS] Строка {row}: ЗАПОЛНЕНО по СНИЛС")
                found = True
        if not found and rec_fio and rec_fio in fio_fallback_index:
            for row in fio_fallback_index[rec_fio]:
                existing = str(ws.cell(row=row, column=COL_DATE).value or "").strip()
                if not existing or existing == 'None':
                    ws.cell(row=row, column=COL_DATE).value = rec_date
                    filled += 1
                    print(f"[FILL-SNILS] Строка {row} (запасной вариант): ЗАПОЛНЕНО")
                    found = True
                    break
    print(f"[FILL-SNILS] ИТОГО заполнено дат: {filled}")
    return filled

# ====================== РџРћРРЎРљ РРќРќ Р’ Р“Р РЈРџРџР• ======================

async def find_inn_in_group(client, group_id, address, topic_id=None):
    try:
        entity = await client.get_entity(int(group_id))
        search_parts = address.split(',')
        search_terms = []
        for part in search_parts:
            part = part.strip()
            if part and len(part) > 3:
                search_terms.append(part)
        async for msg in client.iter_messages(entity, limit=100):
            if not msg.text:
                continue
            msg_lower = msg.text.lower()
            address_lower = address.lower()
            match_count = 0
            for term in search_terms[:3]:
                if term.lower() in msg_lower:
                    match_count += 1
            if match_count >= 2:
                inn = extract_inn_from_text(msg.text)
                if inn:
                    return inn, True
                else:
                    return None, True
        return None, False
    except Exception as e:
        print(f"[GROUP] РћС€РёР±РєР° РїРѕРёСЃРєР°: {e}")
        return None, False

# ====================== РџР•Р Р•РРњР•РќРћР’РђРќРР• Р¤РђР™Р›РћР’ ======================

async def rename_files_by_address(ws, client, group_id, topic_id, add_log, tables_names):
    if not group_id:
        add_log("[РџР•Р Р•РРњР•РќРћР’РђРќРР•] Р“СЂСѓРїРїР° РЅРµ СѓРєР°Р·Р°РЅР° - РїСЂРѕРїСѓСЃРєР°СЋ")
        return tables_names
    add_log("[РџР•Р Р•РРњР•РќРћР’РђРќРР•] РќР°С‡РёРЅР°СЋ РїРµСЂРµРёРјРµРЅРѕРІР°РЅРёРµ РїРѕ Р°РґСЂРµСЃР°Рј...")
    address_map = {}
    table_nums = {}
    for row in range(2, ws.max_row + 1):
        table_num = str(ws.cell(row=row, column=COL_NO).value or "").strip()
        addr = str(ws.cell(row=row, column=COL_ADDR).value or "").strip()
        if table_num and addr and addr != 'None':
            if table_num not in address_map:
                address_map[table_num] = addr
                table_nums[table_num] = row
    add_log(f"[РџР•Р Р•РРњР•РќРћР’РђРќРР•] РќР°Р№РґРµРЅРѕ {len(address_map)} СѓРЅРёРєР°Р»СЊРЅС‹С… Р°РґСЂРµСЃРѕРІ")
    new_names = {}
    for table_num, addr in address_map.items():
        clean_addr = re.sub(r',?\s*РєРІ\.?\s*\d+', '', addr).strip()
        clean_addr = re.sub(r',\s*,', ',', clean_addr)
        add_log(f"[РџР•Р Р•РРњР•РќРћР’РђРќРР•] РС‰Сѓ РРќРќ РґР»СЏ: {clean_addr}")
        inn, found = await find_inn_in_group(client, group_id, clean_addr, topic_id)
        if found and inn:
            new_name = f"{inn}_{clean_addr}"
            add_log(f"[РџР•Р Р•РРњР•РќРћР’РђРќРР•] РќР°Р№РґРµРЅ РРќРќ: {inn} -> {new_name}")
        else:
            new_name = f"РЈРљ_{clean_addr}"
            add_log(f"[РџР•Р Р•РРњР•РќРћР’РђРќРР•] РРќРќ РЅРµ РЅР°Р№РґРµРЅ -> {new_name}")
        new_name = re.sub(r'[<>:"/\\|?*]', '_', new_name)
        new_names[table_num] = new_name
    updated_names = []
    for i, name in enumerate(tables_names):
        table_num = str(i + 1)
        if table_num in new_names:
            updated_names.append(new_names[table_num])
        else:
            updated_names.append(name)
    add_log(f"[РџР•Р Р•РРњР•РќРћР’РђРќРР•] РџРµСЂРµРёРјРµРЅРѕРІР°РЅРѕ {len(new_names)} С„Р°Р№Р»РѕРІ")
    return updated_names

# ====================== РћРџРўРРњРР—РР РћР’РђРќРќР«Р™ РџРћР›РќР«Р™ Р¦РРљР› (v17) ======================

# РљР›Р®Р§Р•Р’Р«Р• РћРџРўРРњРР—РђР¦РР:

# 1. Р¤РёР»СЊС‚СЂ РґР°С‚ РЎРђРњР«Р™ РџР•Р Р’Р«Р™ вЂ” РІС‹СЂРµР·Р°РµРј Р»РёС€РЅРёРµ СЃС‚СЂРѕРєРё РґРѕ РІСЃРµР№ СЂР°Р±РѕС‚С‹

# 2. РљСЌС€ СЃС‚СЂРѕРє РІ РїР°РјСЏС‚Рё вЂ” 1 СЃР±РѕСЂ РІРјРµСЃС‚Рѕ 7 РїРµСЂРµСЃРєР°РЅРёСЂРѕРІР°РЅРёР№

# 3. Batch DeepSeek = 200 (Р±С‹Р»Рѕ 40), Р±РµР· sleep РјРµР¶РґСѓ Р±Р°С‚С‡Р°РјРё, РїР°СЂР°Р»Р»РµР»СЊРЅС‹Рµ РІС‹Р·РѕРІС‹

# 4. РџР°СЂР°Р»Р»РµР»СЊРЅС‹Рµ СЃС‚Р°РґРёРё Р±РѕС‚1+Р±РѕС‚2 (РЎРќРР›РЎ, Р¤РРћ+РґР°С‚Р°)

# 5. РЎРѕС…СЂР°РЅРµРЅРёР№: 4 РІРјРµСЃС‚Рѕ 13

# 6. Sleep'С‹ СѓСЂРµР·Р°РЅС‹ РІ 2-3 СЂР°Р·Р°

# 7. РќРѕСЂРјР°Р»РёР·Р°С†РёСЏ Р°РґСЂРµСЃРѕРІ вЂ” РІ РљРћРќР¦Р• (РїРµСЂРµРґ С‡РµРєРµСЂРѕРј), СЃ РІС‹Р±РѕСЂРѕРј

async def run_full_cycle(ss, bot1, bot2, bot_token, chat_id,
                         items_no_date, items_no_phone, items_snils,
                         year_range, original_rows, tables_names=None, topic_id=None, group_id=None,
                         normalize_addresses=True):
    global stop_requested
    stop_requested = False
    client = await get_client(ss)
    log = []
    t_start = time.time()
    def add(msg):
        ts = datetime.now().strftime('%H:%M:%S')
        log.append(f"[{ts}] {msg}")
        print(f"[LOG] {msg}")
    def elapsed():
        return f"({time.time() - t_start:.1f}СЃ)"
    async def send_status(stage_label):
        if bot_token and chat_id:
            wb.save(result_file)
            await send_file_to_bot(bot_token, chat_id, result_file, f"РўР°Р±Р»РёС†Р° РїРѕСЃР»Рµ: {stage_label}", topic_id)
    async def send_final_zip(complete_session=False):
        if not bot_token or not chat_id:
            return
        split_result = split_by_table_num(
            [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)],
            [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
        )
        if not split_result:
            await send_file_to_bot(bot_token, chat_id, result_file, "РРўРћР“РћР’Р«Р™ Р¤РђР™Р› (РІСЃРµ СЌС‚Р°РїС‹)", topic_id)
            return
        final_names = await rename_files_by_address(ws, client, group_id, topic_id, add, tables_names)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            phones_all = set()
            for table_num, table_data in split_result.items():
                try:
                    geo_result = geo_filter(table_data['headers'], table_data['rows'])
                    phones_all.update(geo_result['phones'])
                    ws_data = [geo_result['headers']] + geo_result['rows']
                    wb_temp = Workbook()
                    ws_temp = wb_temp.active
                    for row in ws_data:
                        ws_temp.append(row)
                    xlsx_buffer = io.BytesIO()
                    wb_temp.save(xlsx_buffer)
                    xlsx_buffer.seek(0)
                    idx = int(table_num) - 1
                    if idx < len(final_names):
                        name = f"{final_names[idx]}.xlsx"
                    else:
                        name = f"Р“Р•Рћ_{table_num}.xlsx"
                    zf.writestr(name, xlsx_buffer.getvalue())
                except Exception as e:
                    add(f"[ZIP] РћС€РёР±РєР° С‚Р°Р±Р»РёС†С‹ {table_num}: {e}")
                    continue
            if phones_all:
                zf.writestr('numbers.txt', '\n'.join(sorted(phones_all)))
            else:
                zf.writestr('numbers.txt', '(РЅРµС‚ РІР°Р»РёРґРЅС‹С… РЅРѕРјРµСЂРѕРІ)')
        zip_buffer.seek(0)
        zip_path = os.path.join(TEMP_DIR, f"result_{int(time.time())}.zip")
        with open(zip_path, 'wb') as f:
            f.write(zip_buffer.getvalue())
        caption = "РРўРћР“РћР’Р«Р™ ZIP РђР РҐРР’ (РІСЃРµ С‚Р°Р±Р»РёС†С‹ + numbers.txt)"
        if complete_session:
            caption = "РЎР•РЎРЎРРЇ Р—РђР’Р•Р РЁР•РќРђ! Р’СЃРµ С„Р°Р№Р»С‹ СЃ РёРјРµРЅР°РјРё РРќРќ/РЈРљ_РђРґСЂРµСЃ."
        await send_zip_to_bot(bot_token, chat_id, zip_path, caption, topic_id)
        add(f"[v] ZIP Р°СЂС…РёРІ РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚ {elapsed()}")
        return zip_path
    async def send_txt_for_max_check():
        if not bot_token or not chat_id:
            return
        phones = []
        for row in range(2, ws.max_row + 1):
            phone_val = str(ws.cell(row=row, column=COL_PHONE).value or "").strip()
            if phone_val and phone_val != 'None' and phone_val != '0':
                clean = clean_phone_without_plus(phone_val)
                if clean:
                    # РЈРґР°Р»СЏРµРј РІРµРґСѓС‰СѓСЋ 7 РґР»СЏ С‡РµРєР° РјР°РєСЃРѕРІ
                    if clean.startswith('7') and len(clean) == 11:
                        clean = clean[1:]
                    phones.append(clean)
        if phones:
            phones = list(set(phones))
            content = '\n'.join(phones)
            caption = "Р§Р•Рљ РњРђРљРЎРћР’ вЂ” РїСЂРѕРІРµСЂСЊС‚Рµ СЌС‚Рё РЅРѕРјРµСЂР° (Р±РµР· +7).\n\nР”РђР›Р•Р•: РѕС‚РїСЂР°РІСЊС‚Рµ TXT СЃ С„РѕСЂРјР°С‚РѕРј 'РќРѕРјРµСЂ РРјСЏ' РґР»СЏ РґРѕР±РёРІР° Р¤РРћ, РёР»Рё РЅР°Р¶РјРёС‚Рµ 'Р—Р°РІРµСЂС€РёС‚СЊ СЃРµСЃСЃРёСЋ' РЅР° СЃР°Р№С‚Рµ."
            await send_txt_to_bot(bot_token, chat_id, content, "check_max.txt", caption, topic_id)
            add(f"[TXT] РћС‚РїСЂР°РІР»РµРЅ check_max.txt СЃ {len(phones)} РЅРѕРјРµСЂР°РјРё (Р±РµР· РІРµРґСѓС‰РµР№ 7) {elapsed()}")
    # ==================== РЎРћР—Р”РђРќРР• РўРђР‘Р›РР¦Р« + Р¤РР›Р¬РўР  Р”РђРў РЎР РђР—РЈ ====================
    result_file = os.path.join(TEMP_DIR, f"result_{int(time.time())}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["N С‚Р°Р±Р»РёС†С‹", "Р¤РРћ", "Р”Р°С‚Р°", "РќРѕРјРµСЂ", "РЎРќРР›РЎ", "РђРґСЂРµСЃСЃ"])
    # РџР°СЂСЃРёРј РґРёР°РїР°Р·РѕРЅ РіРѕРґРѕРІ
    yf, yt = 1945, 1975
    if year_range:
        try:
            parts = year_range.split('-')
            yf, yt = int(parts[0]), int(parts[1])
        except:
            pass
    # Р¤РР›Р¬РўР  Р”РђРў РќРђ РЎРўРђР РўР• вЂ” РІС‹СЂРµР·Р°РµРј СЃС‚СЂРѕРєРё СЃ РЅРµРїРѕРґС…РѕРґСЏС‰РёРј РіРѕРґРѕРј Р Р±РµР· РґР°С‚С‹ Р”Рћ РІСЃРµР№ СЂР°Р±РѕС‚С‹
    rows_kept = 0
    rows_filtered = 0
    rows_no_date_filtered = 0
    years_seen = {}
    for i, row in enumerate(original_rows):
        date_val = str(row[2] if len(row) > 2 else "").strip()  # COL_DATE = 2
        # Р•СЃР»Рё РґР°С‚С‹ РЅРµС‚ СЃРѕРІСЃРµРј вЂ” РІ РјСѓСЃРѕСЂ (Stage 11: РЈРґР°Р»РµРЅРёРµ СЃС‚СЂРѕРє Р±РµР· РґР°С‚С‹)
        if not date_val or date_val == 'None' or date_val == '0':
            rows_no_date_filtered += 1
            continue
        # Р•СЃР»Рё РґР°С‚Р° РµСЃС‚СЊ вЂ” РїСЂРѕРІРµСЂСЏРµРј РіРѕРґ
        parts_d = date_val.split('.')
        if len(parts_d) == 3:
            try:
                year = int(parts_d[2])
                years_seen[year] = years_seen.get(year, 0) + 1
                if year < yf or year > yt:
                    rows_filtered += 1
                    continue
            except ValueError:
                pass  # РќРµРІР°Р»РёРґРЅР°СЏ РґР°С‚Р° вЂ” РЅРµ С„РёР»СЊС‚СЂСѓРµРј, РѕСЃС‚Р°РІР»СЏРµРј
        ws.append([
            str(row[0] if len(row) > 0 else "").strip(),
            str(row[1] if len(row) > 1 else "").strip(),
            date_val,
            str(row[3] if len(row) > 3 else "").strip(),
            str(row[4] if len(row) > 4 else "").strip(),
            str(row[5] if len(row) > 5 else "").strip()
        ])
        rows_kept += 1
    wb.save(result_file)
    add(f"=== Р¤РР›Р¬РўР  Р”РђРў {yf}-{yt} + СѓРґР°Р»РµРЅРёРµ СЃС‚СЂРѕРє Р±РµР· РґР°С‚ ===")
    add(f"  Р’СЃРµРіРѕ СЃС‚СЂРѕРє: {len(original_rows)}")
    add(f"  РћСЃС‚Р°РІР»РµРЅРѕ:   {rows_kept}")
    add(f"  Р’С‹СЂРµР·Р°РЅРѕ РїРѕ РіРѕРґСѓ:    {rows_filtered} (РіРѕРґ РІРЅРµ [{yf}, {yt}])")
    add(f"  Р’С‹СЂРµР·Р°РЅРѕ Р±РµР· РґР°С‚С‹:   {rows_no_date_filtered}")
    if years_seen:
        yr_list = sorted(years_seen.items())
        add(f"  Р“РѕРґР° РІ РґР°РЅРЅС‹С…: {', '.join(f'{y}({c})' for y, c in yr_list[:15])}{'...' if len(yr_list) > 15 else ''}")
    add(f"  {elapsed()}")
    # ==================== РљР­РЁ РЎРўР РћРљ (1 СЃР±РѕСЂ РІРјРµСЃС‚Рѕ 7 РїРµСЂРµСЃРєР°РЅРѕРІ) ====================
    class RowCache:
        """Р•РґРёРЅС‹Р№ РєСЌС€ РІСЃРµС… СЃС‚СЂРѕРє С‚Р°Р±Р»РёС†С‹. РћР±РЅРѕРІР»СЏРµС‚СЃСЏ РїСЂРё РёР·РјРµРЅРµРЅРёСЏС…."""
        def __init__(self):
            self.rows = {}  # row_number -> {fio, date, phone, snils, addr, table_num}
            self.rebuild()
        def rebuild(self):
            self.rows.clear()
            self._no_date = None
            self._no_phone = None
            self._snils_no_date = None
            self._phones_no_date = None
            self._addrs_no_apt = None
            self._addrs_with_apt = None
            for row in range(2, ws.max_row + 1):
                self.rows[row] = {
                    'fio': str(ws.cell(row=row, column=COL_FIO).value or "").strip(),
                    'date': str(ws.cell(row=row, column=COL_DATE).value or "").strip(),
                    'phone': str(ws.cell(row=row, column=COL_PHONE).value or "").strip(),
                    'snils': str(ws.cell(row=row, column=COL_SNILS).value or "").strip(),
                    'addr': str(ws.cell(row=row, column=COL_ADDR).value or "").strip(),
                    'table_num': str(ws.cell(row=row, column=COL_NO).value or "").strip(),
                }
        def invalidate(self):
            self._no_date = None
            self._no_phone = None
            self._snils_no_date = None
            self._phones_no_date = None
            self._addrs_no_apt = None
            self._addrs_with_apt = None
        @property
        def no_date_items(self):
            if self._no_date is None:
                self._no_date = [
                    (r, d) for r, d in self.rows.items()
                    if d['fio'] and d['fio'] != 'None'
                    and d['phone'] and d['phone'] != 'None' and d['phone'] != '0'
                    and (not d['date'] or d['date'] == 'None' or d['date'] == '0')
                ]
            return self._no_date
        @property
        def no_phone_items(self):
            if self._no_phone is None:
                self._no_phone = [
                    (r, d) for r, d in self.rows.items()
                    if d['fio'] and d['fio'] != 'None'
                    and d['date'] and d['date'] != 'None' and d['date'] != '0'
                    and (not d['phone'] or d['phone'] == 'None' or d['phone'] == '0')
                ]
            return self._no_phone
        @property
        def snils_no_date(self):
            if self._snils_no_date is None:
                seen = set()
                result = []
                for r, d in self.rows.items():
                    if (not d['date'] or d['date'] == 'None' or d['date'] == '0'):
                        snils = clean_snils(d['snils'])
                        if snils and len(snils) >= 11 and snils not in seen:
                            seen.add(snils)
                            result.append(snils)
                self._snils_no_date = result
            return self._snils_no_date
        @property
        def phones_no_date(self):
            if self._phones_no_date is None:
                seen = set()
                result = []
                for r, d in self.rows.items():
                    if not d['date'] or d['date'] == 'None' or d['date'] == '0':
                        phone = clean_phone(d['phone'])
                        if phone and phone != 'None' and phone != '0' and phone not in seen:
                            seen.add(phone)
                            result.append((r, phone))
                self._phones_no_date = result
            return self._phones_no_date
        @property
        def addrs_without_apt(self):
            if self._addrs_no_apt is None:
                self._addrs_no_apt = []
                self._addrs_with_apt = []
                seen_no = set()
                seen_with = set()
                for r, d in self.rows.items():
                    addr = d['addr']
                    if not addr or addr == 'None':
                        continue
                    has_apt = bool(re.search(r',\s*\d+\s*$', addr))
                    norm = normalize_address_local(addr)
                    if has_apt:
                        if norm not in seen_with:
                            seen_with.add(norm)
                            self._addrs_with_apt.append(norm)
                    else:
                        phone = clean_phone(d['phone'])
                        if phone and phone != 'None' and phone != '0':
                            if norm not in seen_no:
                                seen_no.add(norm)
                                self._addrs_no_apt.append({
                                    'row': r, 'address': addr,
                                    'address_normalized': norm, 'phone': phone
                                })
            return self._addrs_no_apt
    cache = RowCache()
    # ==================== РќРћР РњРђР›РР—РђР¦РРЇ Р¤РРћ (DeepSeek, batch=200) ====================
    add(f"=== DeepSeek: РЅРѕСЂРјР°Р»РёР·Р°С†РёСЏ Р¤РРћ ===")
    fio_rows = []
    fio_values = []
    for r, d in cache.rows.items():
        if d['fio'] and d['fio'] != 'None':
            fio_rows.append(r)
            fio_values.append(d['fio'])
    add(f"  Р¤РРћ Рє РѕР±СЂР°Р±РѕС‚РєРµ: {len(fio_values)}")
    BATCH = 200  # РЈРІРµР»РёС‡РµРЅРѕ СЃ 40
    fio_changed = 0
    # РџР°СЂР°Р»Р»РµР»СЊРЅС‹Рµ РІС‹Р·РѕРІС‹ DeepSeek РґР»СЏ Р¤РРћ (РїРѕ 3 Р±Р°С‚С‡Р° РѕРґРЅРѕРІСЂРµРјРµРЅРЅРѕ)
    async def process_fio_batch(batch_vals, batch_rows):
        return await normalize_batch_deepseek(batch_vals, 'fio')
    tasks = []
    for batch_idx in range(0, len(fio_values), BATCH):
        batch_vals = fio_values[batch_idx:batch_idx + BATCH]
        batch_rows = fio_rows[batch_idx:batch_idx + BATCH]
        tasks.append((batch_vals, batch_rows))
    # Р“СЂСѓРїРїРёСЂСѓРµРј РїРѕ 3 РїР°СЂР°Р»Р»РµР»СЊРЅС‹С… РІС‹Р·РѕРІР°
    for chunk_start in range(0, len(tasks), 3):
        chunk = tasks[chunk_start:chunk_start + 3]
        chunk_tasks = [process_fio_batch(vals, rows) for vals, rows in chunk]
        results = await asyncio.gather(*chunk_tasks)
        for (vals, rows), normalized in zip(chunk, results):
            for j, norm_val in enumerate(normalized):
                if j < len(rows):
                    old_val = str(ws.cell(row=rows[j], column=COL_FIO).value or "").strip()
                    if old_val != norm_val and norm_val:
                        fio_changed += 1
                    ws.cell(row=rows[j], column=COL_FIO).value = norm_val
        pkg_from = chunk_start // 3 + 1
        pkg_to = min((chunk_start + len(chunk) * 3 - 1) // 3 + 1, len(tasks))
        add(f"  [Р¤РРћ] РџР°РєРµС‚С‹ {pkg_from}-{pkg_to}/{len(tasks)} | РёР·РјРµРЅРµРЅРѕ: {fio_changed} {elapsed()}")
    cache.rebuild()
    wb.save(result_file)
    add(f"  РРўРћР“Рћ Р¤РРћ: РѕР±СЂР°Р±РѕС‚Р°РЅРѕ {len(fio_values)}, РёР·РјРµРЅРµРЅРѕ {fio_changed} {elapsed()}")
    # ==================== РЎРўРђРўРЈРЎ РџР•Р Р•Р” РџР РћР‘РР’РћРњ ====================
    add(f"=== РЎРўРђРўРЈРЎ РџР•Р Р•Р” РџР РћР‘РР’РћРњ ===")
    add(f"  Р’СЃРµРіРѕ СЃС‚СЂРѕРє:     {ws.max_row - 1}")
    add(f"  Р‘РµР· РґР°С‚С‹:         {len(cache.no_date_items)}")
    add(f"  Р‘РµР· РЅРѕРјРµСЂР°:       {len(cache.no_phone_items)}")
    add(f"  РЎРќРР›РЎ Р±РµР· РґР°С‚С‹:   {len(cache.snils_no_date)}")
    add(f"  Р‘РµР· РґР°С‚С‹ (РЅРѕРјРµСЂ): {len(cache.phones_no_date)}")
    add(f"  {elapsed()}")
    # ============ Р­РўРђРџ 1: Р¤РРћ+РќРћРњР•Р  -> Р‘РћРў1 ============
    if cache.no_date_items and not stop_requested:
        items = [(r, d) for r, d in cache.no_date_items]
        cid = f"s1_{int(time.time())}"
        add(f"=== Р­РўРђРџ 1: Р¤РРћ+РќРћРњР•Р  -> Р±РѕС‚1 ===")
        add(f"  РЎС‚СЂРѕРє Рє РїСЂРѕР±РёРІСѓ: {len(items)}")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 1: Р¤РРћ+РќРћРњР•Р ", len(items), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("  [v] РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            txt = "\n".join([f"{normalize_fio_local(d['fio'])}\t{clean_phone(d['phone'])}" for _, d in items])
            tpath = os.path.join(TEMP_DIR, f"t1_{int(time.time())}.txt")
            with open(tpath, 'w', encoding='utf-8') as f:
                f.write(txt)
            add(f"  TXT: {len(items)} СЃС‚СЂРѕРє, {len(txt)} Р±Р°Р№С‚")
            await clear_bot(client, bot1)
            e = await client.get_entity(bot1)
            await client.send_message(e, "РџСЂРѕР±РёРІС‹")
            await asyncio.sleep(1)
            await click_btn(client, bot1, "Р¤РРћ+РЅРѕРјРµСЂ")
            await asyncio.sleep(1)
            last_msgs = await client.get_messages(e, limit=1)
            last_msg_id = last_msgs[0].id if last_msgs else 0
            await client.send_file(e, tpath)
            add("  Р¤Р°Р№Р» РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚1, РѕР¶РёРґР°РЅРёРµ XLSX...")
            before_empty = len(cache.no_date_items)
            msg = await wait_xlsx(client, bot1, 180, since_msg_id=last_msg_id)
            if msg:
                rpath = os.path.join(TEMP_DIR, f"r1_{int(time.time())}.xlsx")
                await client.download_media(msg, file=rpath)
                recs = parse_xlsx(rpath)
                filled = fill_dates_from_response(ws, recs)
                wb.save(result_file)
                cache.rebuild()
                after_empty = len(cache.no_date_items)
                add(f"  РћС‚РІРµС‚РѕРІ: {len(recs)} | Р—Р°РїРѕР»РЅРµРЅРѕ РґР°С‚: {filled} | РћСЃС‚Р°Р»РѕСЃСЊ Р±РµР· РґР°С‚С‹: {after_empty} (Р±С‹Р»Рѕ {before_empty}) {elapsed()}")
                await send_status("СЌС‚Р°Рї 1")
            else:
                add("  [!] Р‘РѕС‚1 РЅРµ РѕС‚РІРµС‚РёР»")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ Р­РўРђРџ 2: РЎРќРР›РЎ -> Р‘РћРў1 ============
    snils_list = cache.snils_no_date
    if snils_list and not stop_requested:
        cid = f"s2_{int(time.time())}"
        add(f"=== Р­РўРђРџ 2: РЎРќРР›РЎ -> Р±РѕС‚1 ===")
        add(f"  РЈРЅРёРєР°Р»СЊРЅС‹С… РЎРќРР›РЎ: {len(snils_list)}")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 2: РЎРќРР›РЎ (Р±РѕС‚1)", len(snils_list), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("  [v] РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            txt = "\n".join(snils_list)
            tpath = os.path.join(TEMP_DIR, f"t2_{int(time.time())}.txt")
            with open(tpath, 'w', encoding='utf-8') as f:
                f.write(txt)
            before_empty = len(cache.snils_no_date)
            await clear_bot(client, bot1)
            e1 = await client.get_entity(bot1)
            await client.send_message(e1, "РџСЂРѕР±РёРІС‹")
            await asyncio.sleep(1)
            await click_btn(client, bot1, "РЎРќРР›РЎ")
            await asyncio.sleep(1)
            last_msgs = await client.get_messages(e1, limit=1)
            last_msg_id = last_msgs[0].id if last_msgs else 0
            await client.send_file(e1, tpath)
            add(f"  TXT РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚1 ({len(snils_list)} РЎРќРР›РЎ), РѕР¶РёРґР°РЅРёРµ XLSX...")
            msg = await wait_xlsx(client, bot1, 180, since_msg_id=last_msg_id)
            if msg:
                rpath = os.path.join(TEMP_DIR, f"r2_{int(time.time())}.xlsx")
                await client.download_media(msg, file=rpath)
                recs = parse_xlsx(rpath)
                filled = fill_snils_dates(ws, recs)
                wb.save(result_file)
                cache.rebuild()
                after_empty = len(cache.snils_no_date)
                add(f"  РћС‚РІРµС‚РѕРІ: {len(recs)} | Р—Р°РїРѕР»РЅРµРЅРѕ РґР°С‚: {filled} | РЎРќРР›РЎ Р±РµР· РґР°С‚С‹: Р±С‹Р»Рѕ {before_empty} -> СЃС‚Р°Р»Рѕ {after_empty} {elapsed()}")
                await send_status("СЌС‚Р°Рї 2")
            else:
                add("  [!] Р‘РѕС‚1 РЅРµ РѕС‚РІРµС‚РёР»")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ Р­РўРђРџ 3: РћРЎРўРђР’РЁРР•РЎРЇ РЎРќРР›РЎ -> Р‘РћРў2 (РґРѕР±РёРІРєР°) ============
    snils_remaining = cache.snils_no_date
    if snils_remaining and not stop_requested:
        cid = f"s3_{int(time.time())}"
        add(f"=== Р­РўРђРџ 3: РЎРќРР›РЎ (РґРѕР±РёРІРєР°) -> Р±РѕС‚2 ===")
        add(f"  РћСЃС‚Р°Р»РѕСЃСЊ РЎРќРР›РЎ Р±РµР· РґР°С‚С‹: {len(snils_remaining)}")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 3: РЎРќРР›РЎ РґРѕР±РёРІРєР° (Р±РѕС‚2)", len(snils_remaining), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("  [v] РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            txt = "\n".join(snils_remaining)
            tpath = os.path.join(TEMP_DIR, f"t3_{int(time.time())}.txt")
            with open(tpath, 'w', encoding='utf-8') as f:
                f.write(txt)
            before_empty = len(cache.snils_no_date)
            e2 = await client.get_entity(bot2)
            last_msgs = await client.get_messages(e2, limit=1)
            last_msg_id = last_msgs[0].id if last_msgs else 0
            await client.send_file(e2, tpath)
            add(f"  TXT РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚2 ({len(snils_remaining)} РЎРќРР›РЎ), РѕР¶РёРґР°РЅРёРµ XLSX...")
            msg = await wait_xlsx(client, bot2, 180, since_msg_id=last_msg_id)
            if msg:
                rpath = os.path.join(TEMP_DIR, f"r3_{int(time.time())}.xlsx")
                await client.download_media(msg, file=rpath)
                recs = parse_xlsx(rpath)
                filled = fill_snils_dates(ws, recs)
                wb.save(result_file)
                cache.rebuild()
                after_empty = len(cache.snils_no_date)
                add(f"  РћС‚РІРµС‚РѕРІ: {len(recs)} | Р—Р°РїРѕР»РЅРµРЅРѕ РґР°С‚: {filled} | РЎРќРР›РЎ Р±РµР· РґР°С‚С‹: Р±С‹Р»Рѕ {before_empty} -> СЃС‚Р°Р»Рѕ {after_empty} {elapsed()}")
                await send_status("СЌС‚Р°Рї 3")
            else:
                add("  [!] Р‘РѕС‚2 РЅРµ РѕС‚РІРµС‚РёР»")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ Р­РўРђРџ 3.5: РџР РћР‘РР’ РџРћ РќРћРњР•Р РЈ -> Р‘РћРў2 ============
    phones_for_probev = cache.phones_no_date
    if phones_for_probev and not stop_requested:
        cid = f"s35_{int(time.time())}"
        add(f"=== Р­РўРђРџ 3.5: РџР РћР‘РР’ РџРћ РќРћРњР•Р РЈ -> Р±РѕС‚2 ===")
        add(f"  РќРѕРјРµСЂРѕРІ Р±РµР· РґР°С‚С‹: {len(phones_for_probev)}")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 3.5: РџР РћР‘РР’ РџРћ РќРћРњР•Р РЈ", len(phones_for_probev), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("  [v] РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            txt_lines = []
            for _, phone in phones_for_probev:
                clean = clean_phone_without_plus(phone)
                if clean:
                    txt_lines.append(clean)
            txt_content = '\n'.join(txt_lines)
            tpath = os.path.join(TEMP_DIR, f"t35_{int(time.time())}.txt")
            with open(tpath, 'w', encoding='utf-8') as f:
                f.write(txt_content)
            before_empty = len(cache.phones_no_date)
            e = await client.get_entity(bot2)
            last_msgs = await client.get_messages(e, limit=1)
            last_msg_id = last_msgs[0].id if last_msgs else 0
            await client.send_file(e, tpath)
            add(f"  TXT РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚2 ({len(txt_lines)} РЅРѕРјРµСЂРѕРІ), РѕР¶РёРґР°РЅРёРµ XLSX...")
            msg = await wait_xlsx(client, bot2, 180, since_msg_id=last_msg_id)
            if msg:
                rpath = os.path.join(TEMP_DIR, f"r35_{int(time.time())}.xlsx")
                await client.download_media(msg, file=rpath)
                recs = parse_xlsx(rpath)
                filled = fill_dates_from_response(ws, recs)
                wb.save(result_file)
                cache.rebuild()
                after_empty = len(cache.phones_no_date)
                add(f"  РћС‚РІРµС‚РѕРІ: {len(recs)} | Р—Р°РїРѕР»РЅРµРЅРѕ РґР°С‚: {filled} | РћСЃС‚Р°Р»РѕСЃСЊ Р±РµР· РґР°С‚С‹: {after_empty} (Р±С‹Р»Рѕ {before_empty}) {elapsed()}")
                await send_status("СЌС‚Р°Рї 3.5")
            else:
                add("  [!] Р‘РѕС‚2 РЅРµ РѕС‚РІРµС‚РёР»")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ Р­РўРђРџ 5: Р¤РРћ+Р”РђРўРђ -> Р‘РћРў1 ============
    no_phone_items = [(r, d) for r, d in cache.no_phone_items]
    if no_phone_items and not stop_requested:
        cid = f"s5_{int(time.time())}"
        add(f"=== Р­РўРђРџ 5: Р¤РРћ+Р”РђРўРђ -> Р±РѕС‚1 ===")
        add(f"  РЎС‚СЂРѕРє Р±РµР· РЅРѕРјРµСЂР°: {len(no_phone_items)}")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 5: Р¤РРћ+Р”РђРўРђ (Р±РѕС‚1)", len(no_phone_items), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("  [v] РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            txt = "\n".join([f"{normalize_fio_local(d['fio'])}\t{d['date']}" for _, d in no_phone_items])
            tpath = os.path.join(TEMP_DIR, f"t5_{int(time.time())}.txt")
            with open(tpath, 'w', encoding='utf-8') as f:
                f.write(txt)
            before_empty = len(cache.no_phone_items)
            await clear_bot(client, bot1)
            e1 = await client.get_entity(bot1)
            await client.send_message(e1, "РџСЂРѕР±РёРІС‹")
            await asyncio.sleep(1)
            await click_btn(client, bot1, "Р¤РРћ+РґР°С‚Р°")
            await asyncio.sleep(1)
            last_msgs = await client.get_messages(e1, limit=1)
            last_msg_id = last_msgs[0].id if last_msgs else 0
            await client.send_file(e1, tpath)
            add(f"  TXT РѕС‚РїСЂР°РІР»РµРЅ РІ Р±РѕС‚1 ({len(no_phone_items)} СЃС‚СЂРѕРє), РѕР¶РёРґР°РЅРёРµ XLSX...")
            msg = await wait_xlsx(client, bot1, 180, since_msg_id=last_msg_id)
            if msg:
                rpath = os.path.join(TEMP_DIR, f"r5_{int(time.time())}.xlsx")
                await client.download_media(msg, file=rpath)
                recs = parse_xlsx(rpath)
                filled = fill_phones_from_response(ws, recs)
                wb.save(result_file)
                cache.rebuild()
                after_empty = len(cache.no_phone_items)
                add(f"  РћС‚РІРµС‚РѕРІ: {len(recs)} | Р—Р°РїРѕР»РЅРµРЅРѕ РЅРѕРјРµСЂРѕРІ: {filled} | Р‘РµР· РЅРѕРјРµСЂР°: Р±С‹Р»Рѕ {before_empty} -> СЃС‚Р°Р»Рѕ {after_empty} {elapsed()}")
                await send_status("СЌС‚Р°Рї 5")
            else:
                add("  [!] Р‘РѕС‚1 РЅРµ РѕС‚РІРµС‚РёР»")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ Р­РўРђРџ 6: РћРЎРўРђР’РЁРР•РЎРЇ Р¤РРћ+Р”РђРўРђ -> Р‘РћРў2 (РґРѕР±РёРІРєР°, РїРѕСЃС‚СЂРѕС‡РЅРѕ РІ TXT) ============
    no_phone_remaining = [(r, d) for r, d in cache.no_phone_items]
    if no_phone_remaining and not stop_requested:
        cid = f"s6_{int(time.time())}"
        add(f"=== Р­РўРђРџ 6: Р¤РРћ+Р”РђРўРђ (РґРѕР±РёРІРєР°) -> Р±РѕС‚2 (РїРѕСЃС‚СЂРѕС‡РЅС‹Р№ TXT) ===")
        add(f"  РћСЃС‚Р°Р»РѕСЃСЊ Р±РµР· РЅРѕРјРµСЂР°: {len(no_phone_remaining)}")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 6: Р¤РРћ+Р”РђРўРђ РґРѕР±РёРІРєР° (Р±РѕС‚2)", len(no_phone_remaining), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("  [v] РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            before_empty = len(cache.no_phone_items)
            # Р¤РѕСЂРјРёСЂСѓРµРј РїР°СЂС‹ (Р¤РРћ, Р”Р°С‚Р°) РґР»СЏ РїРѕСЃС‚СЂРѕС‡РЅРѕРіРѕ TXT
            fio_date_pairs = []
            for _, d in no_phone_remaining:
                fio_norm = normalize_fio_local(d['fio'])
                date_val = d['date']
                if fio_norm and date_val:
                    fio_date_pairs.append((fio_norm, date_val))
            e2 = await client.get_entity(bot2)
            phone_map = await dobiv_fio_date_line_by_line(client, e2, fio_date_pairs, add)
            # Р—Р°РїРѕР»РЅСЏРµРј РЅРѕРјРµСЂР° РІ С‚Р°Р±Р»РёС†Рµ
            phones_filled = 0
            for clean_ph, (fio, date) in phone_map.items():
                for row_num, d in no_phone_remaining:
                    table_fio = normalize_fio_local(d['fio'])
                    table_date = d['date']
                    if normalize_fio_local(fio).lower() in table_fio.lower() and date == table_date:
                        existing = str(ws.cell(row=row_num, column=COL_PHONE).value or "").strip()
                        if not existing or existing == 'None' or existing == '0':
                            ws.cell(row=row_num, column=COL_PHONE).value = clean_ph
                            phones_filled += 1
                        break
            wb.save(result_file)
            cache.rebuild()
            after_empty = len(cache.no_phone_items)
            add(f"  Р—Р°РїРѕР»РЅРµРЅРѕ РЅРѕРјРµСЂРѕРІ: {phones_filled} | Р‘РµР· РЅРѕРјРµСЂР°: Р±С‹Р»Рѕ {before_empty} -> СЃС‚Р°Р»Рѕ {after_empty} {elapsed()}")
            await send_status("СЌС‚Р°Рї 6")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ Р­РўРђРџ 7: Р”РћР‘РР’ РЎРђРЈР РћРќ -> Р‘РћРў2 (РїРѕСЃР»РµРґРѕРІР°С‚РµР»СЊРЅС‹Р№, РєР°Рє РІ СЌС‚Р°Р»РѕРЅРЅРѕРј РєРѕРґРµ) ============
    no_phone_final = [(r, d) for r, d in cache.no_phone_items]
    if no_phone_final and not stop_requested:
        cid = f"s7_{int(time.time())}"
        add(f"=== Р­РўРђРџ 7: Р”РћР‘РР’ РЎРђРЈР РћРќ -> Р±РѕС‚2 (РїРѕСЃР»РµРґРѕРІР°С‚РµР»СЊРЅС‹Р№ РїСЂРѕР±РёРІ) ===")
        add(f"  РЎС‚СЂРѕРє Р±РµР· РЅРѕРјРµСЂР°: {len(no_phone_final)}")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 7: Р”РћР‘РР’ РЎРђРЈР РћРќ", len(no_phone_final), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("  [v] РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            e = await client.get_entity(bot2)
            # Р¤РѕСЂРјРёСЂСѓРµРј СЃРїРёСЃРѕРє СѓРЅРёРєР°Р»СЊРЅС‹С… Р·Р°РїСЂРѕСЃРѕРІ (row_num, fio, date)
            seen = set()
            queries = []
            for row_num, d in no_phone_final:
                fio_norm = normalize_fio_local(d['fio'])
                date_val = d['date']
                key = (fio_norm, date_val)
                if key not in seen and fio_norm and date_val:
                    seen.add(key)
                    queries.append((row_num, fio_norm, date_val))
            add(f"  РЈРЅРёРєР°Р»СЊРЅС‹С… Р·Р°РїСЂРѕСЃРѕРІ: {len(queries)} (РїРѕСЃР»РµРґРѕРІР°С‚РµР»СЊРЅС‹Р№ РїСЂРѕР±РёРІ)")
            phones_filled = await dobiv_sauron_sequential(
                client, e, queries, add, ws, cache, wb, result_file
            )
            wb.save(result_file)
            cache.rebuild()
            add(f"  РРўРћР“Рћ Р”РћР‘РР’ РЎРђРЈР РћРќ: Р·Р°РїРѕР»РЅРµРЅРѕ РЅРѕРјРµСЂРѕРІ {phones_filled} {elapsed()}")
            await send_status("СЌС‚Р°Рї 7")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ Р­РўРђРџ 8: Р”РћР‘РР’ РљР’РђР РўРР  (TXT РїРѕРґС…РѕРґ: РЅРѕРјРµСЂР° в†’ Р±РѕС‚2 в†’ TXT в†’ DeepSeek) ============
    addrs_no_apt = cache.addrs_without_apt
    if addrs_no_apt and not stop_requested:
        cid = f"s8_{int(time.time())}"
        # РЎРѕР±РёСЂР°РµРј Р°РґСЂРµСЃР° СЃ РєРІР°СЂС‚РёСЂР°РјРё РґР»СЏ РїСЂРёРјРµСЂР°
        addrs_with_apt_list = []
        for r, d in cache.rows.items():
            addr = d['addr']
            if addr and addr != 'None' and re.search(r',\s*\d+\s*$', addr):
                norm = normalize_address_local(addr)
                if norm not in addrs_with_apt_list:
                    addrs_with_apt_list.append(norm)
        add(f"=== Р­РўРђРџ 8: Р”РћР‘РР’ РљР’РђР РўРР  ({len(addrs_no_apt)} Р°РґСЂРµСЃРѕРІ Р±РµР· РєРІ, {len(addrs_with_apt_list)} РїСЂРёРјРµСЂРѕРІ) {elapsed()} ===")
        result = await safe_confirm_with_buttons(bot_token, chat_id, "Р­РўРђРџ 8: Р”РћР‘РР’ РљР’РђР РўРР ", len(addrs_no_apt), cid, add, topic_id)
        if result == "stop":
            await send_final_zip()
            return {"ok": True, "log": log, "stopped": True}
        if result == "skip":
            add("[v] Р­РўРђРџ 8 РџР РћРџРЈР©Р•Рќ")
        elif result == "confirm":
            e = await client.get_entity(bot2)
            all_phones = [item['phone'] for item in addrs_no_apt if item['phone']]
            filled_count = await dobiv_apartments_via_txt(
                client, e, all_phones, add, ws, cache, wb, result_file
            )
            wb.save(result_file)
            cache.rebuild()
            await send_status("СЌС‚Р°Рї 8")
            add(f"[РљР’РђР РўРР Р«] Р—Р°РїРѕР»РЅРµРЅРѕ: {filled_count} {elapsed()}")
    if stop_requested:
        await send_final_zip()
        return {"ok": True, "log": log, "stopped": True}
    # ============ РќРћР РњРђР›РР—РђР¦РРЇ РђР”Р Р•РЎРћР’ (Р’ РљРћРќР¦Р•, РµСЃР»Рё РІРєР»СЋС‡РµРЅР°) ============
    if normalize_addresses:
        add(f"DeepSeek: РЅРѕСЂРјР°Р»РёР·Р°С†РёСЏ Р°РґСЂРµСЃРѕРІ (РІ РєРѕРЅС†Рµ)... {elapsed()}")
        addr_rows = []
        addr_values = []
        for r, d in cache.rows.items():
            addr = d['addr']
            if addr and addr != 'None' and len(addr) > 5:
                addr_rows.append(r)
                addr_values.append(addr)
        async def process_addr_batch(batch_vals, batch_rows):
            return await normalize_batch_deepseek(batch_vals, 'address')
        tasks = []
        for batch_idx in range(0, len(addr_values), BATCH):
            batch_vals = addr_values[batch_idx:batch_idx + BATCH]
            batch_rows = addr_rows[batch_idx:batch_idx + BATCH]
            tasks.append((batch_vals, batch_rows))
        for chunk_start in range(0, len(tasks), 3):
            chunk = tasks[chunk_start:chunk_start + 3]
            chunk_tasks = [process_addr_batch(vals, rows) for vals, rows in chunk]
            results = await asyncio.gather(*chunk_tasks)
            for (vals, rows), normalized in zip(chunk, results):
                for j, norm_val in enumerate(normalized):
                    if j < len(rows):
                        ws.cell(row=rows[j], column=COL_ADDR).value = norm_val
        wb.save(result_file)
        cache.rebuild()
        add(f"РђРґСЂРµСЃР° РЅРѕСЂРјР°Р»РёР·РѕРІР°РЅС‹ {elapsed()}")
    # ============ Р¤РРќРђР› ============
    add(f"=== Р¤РРќРђР› ===")
    # РћС‚РїСЂР°РІР»СЏРµРј TXT РґР»СЏ С‡РµРєР° РјР°РєСЃРѕРІ (Р±РµР· РІРµРґСѓС‰РµР№ 7)
    await send_txt_for_max_check()
    # Р¤РёРЅР°Р»СЊРЅС‹Р№ ZIP
    await send_final_zip()
    # РС‚РѕРіРѕРІР°СЏ СЃС‚Р°С‚РёСЃС‚РёРєР°
    total_dates = 0
    total_phones = 0
    total_snils = 0
    total_addrs = 0
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=COL_DATE).value or "").strip() not in ('', 'None', '0'):
            total_dates += 1
        if str(ws.cell(row=row, column=COL_PHONE).value or "").strip() not in ('', 'None', '0'):
            total_phones += 1
        if str(ws.cell(row=row, column=COL_SNILS).value or "").strip() not in ('', 'None', '0'):
            total_snils += 1
        if str(ws.cell(row=row, column=COL_ADDR).value or "").strip() not in ('', 'None'):
            total_addrs += 1
    final_rows = ws.max_row - 1
    date_pct = round(total_dates / final_rows * 100, 1) if final_rows else 0
    phone_pct = round(total_phones / final_rows * 100, 1) if final_rows else 0
    add(f"=== РРўРћР“РћР’РђРЇ РЎРўРђРўРРЎРўРРљРђ ===")
    add(f"  РЎС‚СЂРѕРє РІСЃРµРіРѕ:        {final_rows}")
    add(f"  РЎ РґР°С‚Р°РјРё:           {total_dates} ({date_pct}%)")
    add(f"  РЎ РЅРѕРјРµСЂР°РјРё:         {total_phones} ({phone_pct}%)")
    add(f"  РЎРѕ РЎРќРР›РЎ:           {total_snils}")
    add(f"  РЎ Р°РґСЂРµСЃР°РјРё:         {total_addrs}")
    add(f"  РћР±С‰РµРµ РІСЂРµРјСЏ:        {elapsed()}")
    add(f"=== ГОТОВО ===")
    # Очистка временных файлов старше 1 часа
    _cleanup_old_temp_files(TEMP_DIR, max_age_seconds=3600)
    return {"ok": True, "log": log, "stopped": stop_requested}

# ====================== ОЧИСТКА ВРЕМЕННЫХ ФАЙЛОВ ======================

def _cleanup_old_temp_files(temp_dir, max_age_seconds=3600):
    """Удаляет временные файлы старше max_age_seconds в temp_dir (кроме .gitkeep)."""
    try:
        now = time.time()
        for fname in os.listdir(temp_dir):
            fpath = os.path.join(temp_dir, fname)
            if not os.path.isfile(fpath):
                continue
            if fname == '.gitkeep':
                continue
            try:
                if now - os.path.getmtime(fpath) > max_age_seconds:
                    os.remove(fpath)
            except OSError:
                pass
    except Exception:
        pass

# ====================== Р­РќР”РџРћРРќРўР« ======================

async def handle_health(request):
    return web.json_response({"ok": True, "message": "X Backend v17.0 (РѕРїС‚РёРјРёР·РёСЂРѕРІР°РЅРЅС‹Р№)"})

async def handle_root(request):
    try:
        server_dir = os.path.dirname(os.path.abspath(__file__))
        # Ищем любой .html файл в директории сервера (устойчиво к проблемам кодировки имени)
        html_path = None
        for fname in os.listdir(server_dir):
            if fname.lower().endswith('.html'):
                html_path = os.path.join(server_dir, fname)
                break
        if not html_path or not os.path.isfile(html_path):
            print(f"[ROOT] HTML-файл не найден в {server_dir}")
            return web.Response(
                text="<html><body><h1>X Backend</h1><p>HTML-файл не найден. Проверьте наличие .html в папке сервера.</p></body></html>",
                content_type="text/html"
            )
        with open(html_path, "r", encoding="utf-8") as f:
            html = f.read()
        return web.Response(text=html, content_type="text/html")
    except Exception as e:
        print(f"[ROOT] Ошибка загрузки HTML: {e}")
        traceback.print_exc()
        return web.Response(
            text=f"<html><body><h1>Ошибка сервера</h1><pre>{e}</pre></body></html>",
            content_type="text/html"
        )

async def handle_upload_zip(request):
    try:
        reader = await request.multipart()
        field = await reader.next()
        if field.name != 'file':
            return web.json_response({"ok": False, "error": "РќРµС‚ С„Р°Р№Р»Р°"}, status=400)
        data = await field.read()
        zip_path = os.path.join(TEMP_DIR, f"upload_{int(time.time())}.zip")
        with open(zip_path, 'wb') as f:
            f.write(data)
        tables_data = []
        with zipfile.ZipFile(zip_path, 'r') as zf:
            for file_info in zf.filelist:
                if file_info.filename.endswith('.xlsx'):
                    with zf.open(file_info) as f:
                        xlsx_data = f.read()
                        xlsx_path = os.path.join(TEMP_DIR, f"tmp_{int(time.time())}_{file_info.filename}")
                        with open(xlsx_path, 'wb') as out:
                            out.write(xlsx_data)
                        wb = load_workbook(xlsx_path, data_only=True)
                        ws = wb.active
                        headers = [str(cell.value or "").strip() for cell in ws[1]]
                        rows = []
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if any(cell for cell in row):
                                rows.append([str(cell or "").strip() for cell in row])
                        tables_data.append({'headers': headers, 'rows': rows, 'name': file_info.filename})
                        os.remove(xlsx_path)
        merged = merge_tables(tables_data)
        if not merged:
            return web.json_response({"ok": False, "error": "РўР°Р±Р»РёС†С‹ РЅРµ РЅР°Р№РґРµРЅС‹"}, status=400)
        merged_path = os.path.join(TEMP_DIR, f"merged_{int(time.time())}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(merged['headers'])
        for row in merged['rows']:
            ws.append(row)
        wb.save(merged_path)
        names = [t.get('name', f'РўР°Р±Р»РёС†Р°_{i+1}') for i, t in enumerate(tables_data)]
        return web.json_response({
            "ok": True,
            "file": merged_path,
            "headers": merged['headers'],
            "rows": merged['rows'],
            "count": len(merged['rows']),
            "tables_count": len(tables_data),
            "names": names
        })
    except Exception as e:
        print(f"[UPLOAD] РћС€РёР±РєР°: {e}")
        traceback.print_exc()
        return web.json_response({"ok": False, "error": str(e)}, status=500)

async def handle_send_code(request):
    try:
        d = await request.json()
        phone = d.get("phone", "").strip()
        if not phone:
            return web.json_response({"ok": False, "error": "Р’РІРµРґРёС‚Рµ РЅРѕРјРµСЂ"}, status=400)
        c = TelegramClient(StringSession(), API_ID, API_HASH)
        await c.connect()
        try:
            r = await c.send_code_request(phone)
        except Exception:
            await c.disconnect()
            raise
        sessions[phone] = {"client": c, "hash": r.phone_code_hash}
        print(f"[AUTH] Код отправлен на {phone}")
        return web.json_response({"ok": True})
    except Exception as e:
        print(f"[AUTH] РћС€РёР±РєР°: {e}")
        return web.json_response({"ok": False, "error": str(e)}, status=400)

async def handle_verify_code(request):
    try:
        d = await request.json()
        phone = d.get("phone", "").strip()
        code = d.get("code", "").strip()
        password = d.get("password", "").strip()
        if phone not in sessions:
            return web.json_response({"ok": False, "error": "РЎРЅР°С‡Р°Р»Р° РѕС‚РїСЂР°РІСЊС‚Рµ РєРѕРґ"}, status=400)
        s = sessions[phone]
        c = s["client"]
        try:
            await c.sign_in(phone, code)
        except SessionPasswordNeededError:
            if not password:
                return web.json_response({"ok": True, "need_2fa": True})
            await c.sign_in(password=password)
        ss = c.session.save()
        me = await c.get_me()
        del sessions[phone]
        user_clients[ss] = c
        print(f"[AUTH] Р’С…РѕРґ: {me.first_name}")
        return web.json_response({
            "ok": True,
            "session": ss,
            "phone": phone,
            "name": me.first_name or ""
        })
    except Exception as e:
        print(f"[AUTH] РћС€РёР±РєР°: {e}")
        return web.json_response({"ok": False, "error": str(e)}, status=400)

async def handle_upload_check_txt(request):
    """Р—Р°РіСЂСѓР·РєР° TXT СЃ РЅРѕРјРµСЂР°РјРё РґР»СЏ РґРѕР±РёРІР° (РєР°Рє РІ minecraft.py)"""
    try:
        reader = await request.multipart()
        field = await reader.next()
        if field.name != 'file':
            return web.json_response({"ok": False, "error": "РќРµС‚ С„Р°Р№Р»Р°"}, status=400)
        data = await field.read()
        txt_content = data.decode('utf-8', errors='ignore')
        # РџР°СЂСЃРёРј TXT: РєР°Р¶РґР°СЏ СЃС‚СЂРѕРєР° - РЅРѕРјРµСЂ РёР»Рё "РЅРѕРјРµСЂ РёРјСЏ"
        phone_to_fio = {}
        for line in txt_content.split('\n'):
            line = line.strip()
            if not line:
                continue
            # РџСЂРѕР±СѓРµРј СЂР°Р·Р±РёС‚СЊ РЅР° РЅРѕРјРµСЂ Рё РёРјСЏ
            parts = line.split(maxsplit=1)
            if len(parts) >= 2:
                phone_raw = parts[0]
                fio = parts[1]
                phone_clean = clean_phone_without_plus(phone_raw)
                if phone_clean and fio:
                    phone_to_fio[phone_clean] = fio
            else:
                # РўРѕР»СЊРєРѕ РЅРѕРјРµСЂ
                phone_clean = clean_phone_without_plus(line)
                if phone_clean:
                    phone_to_fio[phone_clean] = ''
        return web.json_response({
            "ok": True,
            "phone_to_fio": phone_to_fio,
            "count": len(phone_to_fio)
        })
    except Exception as e:
        print(f"[UPLOAD-TXT] РћС€РёР±РєР°: {e}")
        traceback.print_exc()
        return web.json_response({"ok": False, "error": str(e)}, status=500)

async def handle_full_probev(request):
    ss = ""
    try:
        d = await request.json()
        ss = d.get("session", "")
        bot1 = d.get("bot1", "@osint_pam_pam_bot")
        bot2 = d.get("bot2", "@proverim123_bot")
        bot_token = d.get("bot_token", "")
        chat_id = d.get("chat_id", "")
        topic_id = d.get("topic_id", None)
        group_id = d.get("group_id", None)
        year_range = d.get("year_range", "") or "1945-1975"
        items_no_date = d.get("items_no_date", [])
        items_no_phone = d.get("items_no_phone", [])
        items_snils = d.get("items_snils", [])
        original_rows = d.get("original_rows", [])
        tables_names = d.get("tables_names", [])
        normalize_addresses = d.get("normalize_addresses", True)  # РќРѕРІС‹Р№ РїР°СЂР°РјРµС‚СЂ
        if not ss:
            return web.json_response({"ok": False, "error": "РќРµС‚ СЃРµСЃСЃРёРё"}, status=400)
        async with probev_lock:
            if ss in active_probevs:
                task = active_probevs[ss]
                if not task.done():
                    return web.json_response({"ok": False, "error": "РџСЂРѕР±РёРІ СѓР¶Рµ РІС‹РїРѕР»РЅСЏРµС‚СЃСЏ"}, status=409)
                else:
                    del active_probevs[ss]
        print(f"\n{'='*60}")
        print(f"[PROBEV] Р—РђРџРЈРЎРљ РџРћР›РќРћР“Рћ Р¦РРљР›Рђ (8 Р­РўРђРџРћР’ + DeepSeek + Р”РћР‘РР’ РљР’РђР РўРР )")
        print(f"[PROBEV] Р‘РѕС‚1: {bot1}, Р‘РѕС‚2: {bot2}")
        print(f"[PROBEV] РЎС‚СЂРѕРє Р±РµР· РґР°С‚С‹: {len(items_no_date)}")
        print(f"[PROBEV] РЎС‚СЂРѕРє Р±РµР· РЅРѕРјРµСЂР°: {len(items_no_phone)}")
        print(f"[PROBEV] Р“СЂСѓРїРїР°: {group_id or 'РќРµС‚'}")
        print(f"[PROBEV] РўРµРјР°: {topic_id or 'РќРµС‚'}")
        print(f"{'='*60}\n")
        async def run_and_cleanup():
            try:
                result = await run_full_cycle(
                    ss, bot1, bot2, bot_token, chat_id,
                    items_no_date, items_no_phone, items_snils,
                    year_range, original_rows, tables_names, topic_id, group_id,
                    normalize_addresses
                )
                return result
            finally:
                async with probev_lock:
                    if ss in active_probevs:
                        del active_probevs[ss]
                print(f"[PROBEV] Очистка завершена для сессии {str(ss)[:10]}...")
        task = asyncio.create_task(run_and_cleanup())
        async with probev_lock:
            active_probevs[ss] = task
        result = await task
        return web.json_response(result)
    except Exception as e:
        traceback.print_exc()
        async with probev_lock:
            if ss and ss in active_probevs:
                del active_probevs[ss]
        return web.json_response({"ok": False, "error": str(e)}, status=400)

async def handle_stop(request):
    global stop_requested
    stop_requested = True
    return web.json_response({"ok": True, "message": "РћСЃС‚Р°РЅРѕРІРєР° Р·Р°РїСЂРѕС€РµРЅР°"})

async def handle_finish_session(request):
    """Р—Р°РІРµСЂС€РµРЅРёРµ СЃРµСЃСЃРёРё: СЃРѕР·РґР°РЅРёРµ ZIP СЃ СЂР°Р·Р±РёС‚С‹РјРё Р±Р°Р·Р°РјРё (РРќРќ_РђРґСЂРµСЃ.xlsx) Рё РѕС‚РїСЂР°РІРєР° РІ Р±РѕС‚"""
    try:
        d = await request.json()
        ss = d.get("session", "")
        tables_names = d.get("tables_names", [])
        bot_token = d.get("bot_token", "")
        chat_id = d.get("chat_id", "")
        group_id = d.get("group_id", None)
        topic_id = d.get("topic_id", None)
        headers = d.get("headers", [])
        rows = d.get("rows", [])
        # РЎРѕР·РґР°С‘Рј С‚Р°Р±Р»РёС†Сѓ РёР· РїРµСЂРµРґР°РЅРЅС‹С… РґР°РЅРЅС‹С…
        wb = Workbook()
        ws = wb.active
        if headers:
            ws.append(headers)
            for row in rows:
                ws.append(row)
        else:
            # Fallback: РёС‰РµРј РїРѕСЃР»РµРґРЅРёР№ result С„Р°Р№Р»
            result_files = []
            for f in os.listdir(TEMP_DIR):
                if f.startswith('result_') and f.endswith('.xlsx'):
                    result_files.append(os.path.join(TEMP_DIR, f))
            if result_files:
                result_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
                wb = load_workbook(result_files[0], data_only=True)
                ws = wb.active
        if ws.max_row <= 1:
            return web.json_response({"ok": False, "error": "РќРµС‚ РґР°РЅРЅС‹С… РґР»СЏ Р°СЂС…РёРІР°С†РёРё"}, status=400)
        # РЎРѕС…СЂР°РЅСЏРµРј РёС‚РѕРіРѕРІСѓСЋ С‚Р°Р±Р»РёС†Сѓ
        final_path = os.path.join(TEMP_DIR, f"session_final_{int(time.time())}.xlsx")
        wb.save(final_path)
        # РџС‹С‚Р°РµРјСЃСЏ РїРµСЂРµРёРјРµРЅРѕРІР°С‚СЊ РїРѕ РРќРќ
        renamed = tables_names or []
        if ss and group_id:
            try:
                client = await get_client(ss)
                renamed = await rename_files_by_address(ws, client, group_id, topic_id, None, tables_names)
            except Exception as e:
                print(f"[FINISH] РћС€РёР±РєР° РїРµСЂРµРёРјРµРЅРѕРІР°РЅРёСЏ: {e}")
        # Р Р°Р·Р±РёРІР°РµРј РїРѕ N С‚Р°Р±Р»РёС†С‹ Рё СЃРѕР·РґР°С‘Рј ZIP
        split_result = split_by_table_num(
            [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)],
            [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)] for r in range(2, ws.max_row + 1)]
        )
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            if split_result:
                for table_num, table_data in split_result.items():
                    try:
                        geo_result = geo_filter(table_data['headers'], table_data['rows'])
                        ws_data = [geo_result['headers']] + geo_result['rows']
                        wb_temp = Workbook()
                        ws_temp = wb_temp.active
                        for row in ws_data:
                            ws_temp.append(row)
                        xlsx_buffer = io.BytesIO()
                        wb_temp.save(xlsx_buffer)
                        xlsx_buffer.seek(0)
                        idx = int(table_num) - 1
                        name = renamed[idx] if idx < len(renamed) else f"Р“Р•Рћ_{table_num}"
                        name = re.sub(r'[<>:"/\\|?*]', '_', name)
                        zf.writestr(f"{name}.xlsx", xlsx_buffer.getvalue())
                    except Exception as e:
                        print(f"[FINISH-ZIP] РћС€РёР±РєР° С‚Р°Р±Р»РёС†С‹ {table_num}: {e}")
            else:
                # РћРґРЅР° С‚Р°Р±Р»РёС†Р° вЂ” РєР»Р°РґС‘Рј РєР°Рє РµСЃС‚СЊ
                xlsx_buffer = io.BytesIO()
                wb.save(xlsx_buffer)
                xlsx_buffer.seek(0)
                name = renamed[0] if renamed else "result"
                name = re.sub(r'[<>:"/\\|?*]', '_', name)
                zf.writestr(f"{name}.xlsx", xlsx_buffer.getvalue())
            # Р”РѕР±Р°РІР»СЏРµРј numbers.txt
            phones_all = set()
            for row in rows:
                phone_val = str(row[3] if len(row) > 3 else "").strip() if headers else ""
                if not phone_val and len(row) >= 4:
                    # РС‰РµРј РєРѕР»РѕРЅРєСѓ РќРѕРјРµСЂ
                    for i, h in enumerate(headers):
                        if 'РЅРѕРјРµСЂ' in str(h).lower() and i < len(row):
                            phone_val = str(row[i]).strip()
                            break
                if phone_val and phone_val not in ('', 'None', '0'):
                    clean = clean_phone_without_plus(phone_val)
                    if clean:
                        phones_all.add(clean)
            if phones_all:
                zf.writestr('numbers.txt', '\n'.join(sorted(phones_all)))
        zip_buffer.seek(0)
        zip_path = os.path.join(TEMP_DIR, f"session_{int(time.time())}.zip")
        with open(zip_path, 'wb') as f:
            f.write(zip_buffer.getvalue())
        # РћС‚РїСЂР°РІР»СЏРµРј РІ Р±РѕС‚ РµСЃР»Рё РЅР°СЃС‚СЂРѕРµРЅ
        if bot_token and chat_id:
            try:
                await send_zip_to_bot(bot_token, chat_id, zip_path, "РЎРµСЃСЃРёСЏ Р·Р°РІРµСЂС€РµРЅР°! ZIP СЃ Р±Р°Р·Р°РјРё (РРќРќ_РђРґСЂРµСЃ / РЈРљ_РђРґСЂРµСЃ)", topic_id)
            except Exception as e:
                print(f"[FINISH] РћС€РёР±РєР° РѕС‚РїСЂР°РІРєРё РІ Р±РѕС‚: {e}")
        return web.json_response({
            "ok": True,
            "message": f"ZIP СЃРѕР·РґР°РЅ: {len(split_result) if split_result else 1} Р±Р°Р·",
            "files_count": len(split_result) if split_result else 1,
            "zip_path": zip_path
        })
    except Exception as e:
        print(f"[FINISH] РћС€РёР±РєР°: {e}")
        traceback.print_exc()
        return web.json_response({"ok": False, "error": str(e)}, status=500)

async def handle_normalize_addresses(request):
    """РќРѕСЂРјР°Р»РёР·Р°С†РёСЏ Р°РґСЂРµСЃРѕРІ С‡РµСЂРµР· DeepSeek (РґР»СЏ РІРєР»Р°РґРєРё В«РђРґСЂРµСЃР°В»)"""
    try:
        d = await request.json()
        addresses = d.get("addresses", [])
        if not addresses:
            return web.json_response({"ok": False, "error": "РќРµС‚ Р°РґСЂРµСЃРѕРІ"}, status=400)
        # РСЃРїРѕР»СЊР·СѓРµРј РѕР±С‰СѓСЋ С„СѓРЅРєС†РёСЋ РЅРѕСЂРјР°Р»РёР·Р°С†РёРё
        normalized = await normalize_batch_deepseek(addresses, 'address')
        return web.json_response({
            "ok": True,
            "normalized": normalized,
            "count": len(normalized)
        })
    except Exception as e:
        print(f"[NORM-ADDR] РћС€РёР±РєР°: {e}")
        traceback.print_exc()
        # Fallback: Р»РѕРєР°Р»СЊРЅР°СЏ РЅРѕСЂРјР°Р»РёР·Р°С†РёСЏ
        normalized = [normalize_address_local(a) for a in addresses]
        return web.json_response({
            "ok": True,
            "normalized": normalized,
            "count": len(normalized),
            "fallback": True
        })

async def handle_download_file(request):
    """РћС‚РґР°С‡Р° ZIP/XLSX С„Р°Р№Р»Р° РґР»СЏ СЃРєР°С‡РёРІР°РЅРёСЏ"""
    filename = request.match_info.get("filename", "")
    filepath = os.path.join(TEMP_DIR, os.path.basename(filename))
    if not os.path.exists(filepath):
        return web.json_response({"ok": False, "error": "Р¤Р°Р№Р» РЅРµ РЅР°Р№РґРµРЅ"}, status=404)
    return web.FileResponse(filepath)

# ====================== РџРћР”РўР’Р•Р Р–Р”Р•РќРРЇ РЎ РЎРђР™РўРђ ======================

async def handle_pending_confirm(request):
    """Р’РѕР·РІСЂР°С‰Р°РµС‚ С‚РµРєСѓС‰РµРµ РѕР¶РёРґР°СЋС‰РµРµ РїРѕРґС‚РІРµСЂР¶РґРµРЅРёРµ (РґР»СЏ РѕРїСЂРѕСЃР° СЃ СЃР°Р№С‚Р°)"""
    if not pending_web_confirms:
        return web.json_response({"ok": True, "pending": None})
    # РќР°С…РѕРґРёРј СЃР°РјРѕРµ СЃС‚Р°СЂРѕРµ РѕР¶РёРґР°СЋС‰РµРµ РїРѕРґС‚РІРµСЂР¶РґРµРЅРёРµ
    oldest = None
    for cid, data in pending_web_confirms.items():
        if data.get("result") is None:  # РµС‰С‘ РЅРµ РѕС‚РІРµС‡РµРЅРѕ
            if oldest is None or data["timestamp"] < oldest[1]["timestamp"]:
                oldest = (cid, data)
    if oldest:
        return web.json_response({
            "ok": True,
            "pending": {
                "confirm_id": oldest[0],
                "stage": oldest[1]["stage"],
                "count": oldest[1]["count"],
                "elapsed": int(time.time() - oldest[1]["timestamp"])
            }
        })
    return web.json_response({"ok": True, "pending": None})

async def handle_confirm_action(request):
    """РћС‚РІРµС‚ РЅР° РїРѕРґС‚РІРµСЂР¶РґРµРЅРёРµ СЃ СЃР°Р№С‚Р°: confirm, skip, stop, again"""
    try:
        d = await request.json()
        confirm_id = d.get("confirm_id", "")
        action = d.get("action", "confirm")  # confirm, skip, stop, again
        if not confirm_id:
            return web.json_response({"ok": False, "error": "РќРµС‚ confirm_id"}, status=400)
        if confirm_id in pending_web_confirms:
            pending_web_confirms[confirm_id]["result"] = action
            return web.json_response({"ok": True, "message": f"Р”РµР№СЃС‚РІРёРµ {action} РїСЂРёРЅСЏС‚Рѕ"})
        # РњРѕР¶РµС‚ Р±С‹С‚СЊ СѓР¶Рµ РІ pending_confirms (Р±РѕС‚)
        if confirm_id in pending_confirms:
            pending_confirms[confirm_id] = action
            return web.json_response({"ok": True, "message": f"Р”РµР№СЃС‚РІРёРµ {action} РїСЂРёРЅСЏС‚Рѕ (Р±РѕС‚)"})
        return web.json_response({"ok": False, "error": "РџРѕРґС‚РІРµСЂР¶РґРµРЅРёРµ РЅРµ РЅР°Р№РґРµРЅРѕ (СѓСЃС‚Р°СЂРµР»Рѕ?)"}, status=404)
    except Exception as e:
        return web.json_response({"ok": False, "error": str(e)}, status=400)

# ====================== РџРђРЈР—Рђ / Р Р•Р—Р®РњР• ======================

async def handle_pause(request):
    global pause_requested
    pause_requested = True
    return web.json_response({"ok": True, "message": "РџР°СѓР·Р° Р·Р°РїСЂРѕС€РµРЅР°"})

async def handle_resume(request):
    global pause_requested
    pause_requested = False
    return web.json_response({"ok": True, "message": "Р’РѕР·РѕР±РЅРѕРІР»РµРЅРѕ"})

# ====================== MAX РљРћР›РћРќРљРђ ======================

async def handle_fill_max(request):
    """Р”РѕР±Р°РІР»СЏРµС‚ СЃС‚РѕР»Р±РµС† MAX РёР· check_max.txt РѕС‚РІРµС‚Р° (РќРѕРјРµСЂ РРјСЏ)"""
    try:
        d = await request.json()
        ss = d.get("session", "")
        phone_names = d.get("phone_names", {})  # {phone: name}
        if not ss or not phone_names:
            return web.json_response({"ok": False, "error": "РќРµС‚ РґР°РЅРЅС‹С…"}, status=400)
        # РС‰РµРј РїРѕСЃР»РµРґРЅРёР№ result С„Р°Р№Р»
        result_files = []
        for f in os.listdir(TEMP_DIR):
            if f.startswith('result_') and f.endswith('.xlsx'):
                result_files.append(os.path.join(TEMP_DIR, f))
        if not result_files:
            return web.json_response({"ok": False, "error": "РќРµС‚ С„Р°Р№Р»Р° СЂРµР·СѓР»СЊС‚Р°С‚Р°"}, status=404)
        result_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        wb = load_workbook(result_files[0], data_only=True)
        ws = wb.active
        # РћРїСЂРµРґРµР»СЏРµРј РєРѕР»РѕРЅРєСѓ РќРѕРјРµСЂ
        phone_col = None
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=1, column=c).value or "").lower().strip()
            if any(k in val for k in ['РЅРѕРјРµСЂ', 'С‚РµР»РµС„РѕРЅ', 'phone', 'С‚РµР»']):
                phone_col = c
                break
        if not phone_col:
            return web.json_response({"ok": False, "error": "РљРѕР»РѕРЅРєР° РќРѕРјРµСЂ РЅРµ РЅР°Р№РґРµРЅР°"}, status=400)
        # Р”РѕР±Р°РІР»СЏРµРј РєРѕР»РѕРЅРєСѓ MAX РїРѕСЃР»Рµ РїРѕСЃР»РµРґРЅРµР№ СЃСѓС‰РµСЃС‚РІСѓСЋС‰РµР№
        max_col = ws.max_column + 1
        ws.cell(row=1, column=max_col).value = "MAX"
        filled = 0
        for row in range(2, ws.max_row + 1):
            phone_val = str(ws.cell(row=row, column=phone_col).value or "").strip()
            clean = clean_phone_without_plus(phone_val)
            # РџСЂРѕР±СѓРµРј СЃ 7 Рё Р±РµР·
            for try_phone in [clean, clean[1:] if clean.startswith('7') else '7'+clean]:
                if try_phone in phone_names:
                    ws.cell(row=row, column=max_col).value = phone_names[try_phone]
                    filled += 1
                    break
        wb.save(result_files[0])
        return web.json_response({
            "ok": True,
            "filled": filled,
            "total": len(phone_names)
        })
    except Exception as e:
        traceback.print_exc()
        return web.json_response({"ok": False, "error": str(e)}, status=500)

# ====================== Р Р•Р“РРЎРўР РђР¦РРЇ / Р’РҐРћР” (Telegram = Р‘Р”) ======================

# РќРёРєР°РєРѕРіРѕ USER_DB! Р’СЃРµ РґР°РЅРЅС‹Рµ вЂ” РІ РіСЂСѓРїРїР°С… Telegram Р°РєРєР°СѓРЅС‚Р°.

# Р›РѕРіРёРЅ/РїР°СЂРѕР»СЊ РёС‰РµС‚СЃСЏ СЃРєР°РЅРёСЂРѕРІР°РЅРёРµРј Р’РЎР•РҐ РіСЂСѓРїРї/РґРёР°Р»РѕРіРѕРІ Р°РєРєР°СѓРЅС‚Р°.

app = web.Application(middlewares=[log_and_cors], client_max_size=200 * 1024 * 1024)
app.router.add_get("/", handle_root)
app.router.add_get("/health", handle_health)
app.router.add_post("/upload-zip", handle_upload_zip)
app.router.add_post("/upload-check-txt", handle_upload_check_txt)
app.router.add_post("/send-code", handle_send_code)
app.router.add_post("/verify-code", handle_verify_code)
app.router.add_post("/full-probev", handle_full_probev)
app.router.add_post("/stop", handle_stop)
app.router.add_post("/finish-session", handle_finish_session)
app.router.add_post("/normalize-addresses", handle_normalize_addresses)
app.router.add_get("/download/{filename}", handle_download_file)

async def on_startup(app):
    port = app["port"]
    msg = (
        "=" * 60 + "\n"
        f"X Backend v17.0 Р—РђРџРЈР©Р•Рќ (Telegram=Р‘Р”: РіСЂСѓРїРїС‹=РїР°РїРєРё, Р±РµР· USER_DB)\n"
        f"Host: 0.0.0.0  |  Port: {port}\n"
        + "=" * 60
    )
    print(msg, flush=True)

async def on_shutdown(app):
    print("[SHUTDOWN] Р—Р°РєСЂС‹РІР°СЋ СЃРѕРµРґРёРЅРµРЅРёСЏ...", flush=True)
    for ss, client in list(user_clients.items()):
        try:
            if client.is_connected():
                await client.disconnect()
        except Exception:
            pass
    user_clients.clear()
    sessions.clear()
    pending_confirms.clear()
    print("[SHUTDOWN] Р“РѕС‚РѕРІРѕ.", flush=True)
if __name__ == "__main__":
    import sys, signal as _signal
    try:
        port = int(os.environ.get("PORT", 4545))
        app["port"] = port
        app.on_startup.append(on_startup)
        app.on_shutdown.append(on_shutdown)
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        runner = web.AppRunner(app)
        loop.run_until_complete(runner.setup())
        import socket as _socket
        sock = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
        sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEADDR, 1)
        try:
            sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEPORT, 1)
        except:
            pass
        sock.bind(("0.0.0.0", port))
        sock.listen(128)
        sock.setblocking(False)
        site = web.SockSite(runner, sock)
        loop.run_until_complete(site.start())
        print(f"[START] РЎР•Р Р’Р•Р  Р“РћРўРћР’ вЂ” РїРѕСЂС‚ {port}", flush=True)
        def shutdown():
            print("[SIGNAL] РћСЃС‚Р°РЅР°РІР»РёРІР°СЋ...", flush=True)
            loop.create_task(_do_shutdown(runner, sock, loop))
        async def _do_shutdown(runner, sock, loop):
            await runner.cleanup()
            sock.close()
            loop.stop()
        for sig in (_signal.SIGTERM, _signal.SIGINT):
            try:
                loop.add_signal_handler(sig, shutdown)
            except NotImplementedError:
                pass
        loop.run_forever()
        loop.close()
    except Exception as e:
        print(f"[FATAL] {e}", flush=True)
        traceback.print_exc()
        sys.exit(1)
